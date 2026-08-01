"""
Regenera supabase/seed_event_clients.json desde el Excel de clientes Carranza 50.

Uso:
  python scripts/import_event_clients_from_excel.py
  python scripts/import_event_clients_from_excel.py --excel "I:\\Mi unidad\\Eventos\\lista de clientes Carranza 50.xlsx"
  python scripts/import_event_clients_from_excel.py --seed   # además upsert a Supabase

Env:
  EVENTOS_CLIENTS_EXCEL  ruta al .xlsx (override del default)
  NEXT_PUBLIC_SUPABASE_URL + SUPABASE_SERVICE_ROLE_KEY en .env.local (para --seed)

No borra clientes existentes en Supabase: solo inserta faltantes (match
company+email / company+phone / company+contact).
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
import urllib.error
import urllib.request
from datetime import datetime, timezone
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
SEED = ROOT / "supabase" / "seed_event_clients.json"
ENV_PATH = ROOT / ".env.local"
DEFAULT_EXCEL = Path(
    os.environ.get(
        "EVENTOS_CLIENTS_EXCEL",
        r"I:\Mi unidad\Eventos\lista de clientes Carranza 50.xlsx",
    )
)

# Reutilizar normalización de teléfono del helper de Seguimiento si está disponible
sys.path.insert(0, str(ROOT / "ingestor"))
try:
    from eventos_seguimiento import normalize_email, normalize_phone  # type: ignore
except Exception:  # pragma: no cover

    def normalize_phone(value):
        if value is None:
            return None
        if isinstance(value, float):
            if value != value:
                return None
            value = int(value) if value == int(value) else value
        digits = re.sub(r"\D", "", str(value))
        if len(digits) == 12 and digits.startswith("52"):
            digits = digits[2:]
        return digits or None

    def normalize_email(value):
        if value is None:
            return None
        s = str(value).strip().lower()
        return s if s and "@" in s else None


def load_env(path: Path) -> dict[str, str]:
    vals: dict[str, str] = {}
    if not path.exists():
        return vals
    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, v = line.split("=", 1)
        vals[k.strip()] = v.strip().strip('"').strip("'")
    return vals


def http_json(method: str, url: str, headers: dict[str, str], body=None):
    data = None if body is None else json.dumps(body).encode("utf-8")
    req = urllib.request.Request(url, data=data, method=method)
    for k, v in headers.items():
        req.add_header(k, v)
    if body is not None:
        req.add_header("Content-Type", "application/json")
        req.add_header("Prefer", "return=minimal")
    try:
        with urllib.request.urlopen(req, timeout=120) as resp:
            text = resp.read().decode("utf-8", errors="replace")
            return resp.status, json.loads(text) if text else None
    except urllib.error.HTTPError as e:
        return e.code, e.read().decode("utf-8", errors="replace")


def _cell(row: tuple, idx: int):
    if idx < 0 or idx >= len(row):
        return None
    return row[idx]


def find_header(rows: list) -> tuple[int, dict[str, int]]:
    aliases = {
        "company": ("nombre de la empresa", "empresa", "razon social", "compañia", "compania"),
        "contact": ("contacto", "nombre", "nombre del contacto"),
        "email": ("correo electronico", "correo", "email", "e-mail"),
        "phone": ("telefono", "teléfono", "celular", "whatsapp"),
    }

    def n(s) -> str:
        if s is None:
            return ""
        t = str(s).strip().lower()
        t = (
            t.replace("á", "a")
            .replace("é", "e")
            .replace("í", "i")
            .replace("ó", "o")
            .replace("ú", "u")
            .replace("ñ", "n")
        )
        return t

    for i, row in enumerate(rows[:40]):
        headers = [n(c) for c in row]
        mapping: dict[str, int] = {}
        for field, names in aliases.items():
            for j, h in enumerate(headers):
                if h in names:
                    mapping[field] = j
                    break
        if "company" in mapping:
            return i, mapping
    raise SystemExit("No se encontró encabezado con columna de empresa en el Excel.")


def read_excel(path: Path) -> list[dict]:
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise SystemExit(
            "Falta openpyxl. Instala: pip install -r ingestor/requirements.txt"
        ) from e

    if not path.exists():
        raise SystemExit(f"Excel no encontrado: {path}")

    wb = load_workbook(path, read_only=True, data_only=True)
    # Prefer hoja Empresas; si no, la primera con datos
    sheet_name = "Empresas" if "Empresas" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    hdr_i, cols = find_header(rows)

    raw: list[dict] = []
    skipped_empty = 0
    for row in rows[hdr_i + 1 :]:
        if not row:
            skipped_empty += 1
            continue
        company = str(_cell(row, cols["company"]) or "").strip()
        if not company:
            skipped_empty += 1
            continue
        contact = str(_cell(row, cols.get("contact", -1)) or "").strip() or None
        email = normalize_email(_cell(row, cols["email"]) if "email" in cols else None)
        phone = normalize_phone(_cell(row, cols["phone"]) if "phone" in cols else None)
        raw.append(
            {
                "company": company,
                "contact": contact,
                "email": email,
                "phone": phone,
            }
        )

    # Dedup: misma empresa + (email|phone|contact) idénticos
    seen: set[tuple] = set()
    out: list[dict] = []
    dupes = 0
    for r in raw:
        key = (
            r["company"].lower(),
            (r["email"] or "").lower(),
            r["phone"] or "",
            (r["contact"] or "").lower(),
        )
        if key in seen:
            dupes += 1
            continue
        seen.add(key)
        out.append(r)

    print(
        f"Excel hoja={sheet_name} filas_utiles={len(raw)} "
        f"tras_dedup={len(out)} vacias={skipped_empty} dupes_exactos={dupes}"
    )
    return out


def client_identity(company: str, contact, email, phone) -> str:
    c = (company or "").strip().lower()
    e = (email or "").strip().lower()
    p = (phone or "").strip()
    n = (contact or "").strip().lower()
    if e:
        return f"{c}|e:{e}"
    if p:
        return f"{c}|p:{p}"
    if n:
        return f"{c}|n:{n}"
    return f"{c}|solo"


def seed_supabase(rows: list[dict]) -> int:
    env = load_env(ENV_PATH)
    url = env.get("NEXT_PUBLIC_SUPABASE_URL", "").rstrip("/")
    key = env.get("SUPABASE_SERVICE_ROLE_KEY", "")
    if not url or not key:
        print("BLOCKED: need NEXT_PUBLIC_SUPABASE_URL + SUPABASE_SERVICE_ROLE_KEY in .env.local")
        return 2

    headers = {"apikey": key, "Authorization": f"Bearer {key}"}
    status, existing = http_json(
        "GET",
        f"{url}/rest/v1/event_clients?select=company_name,contact_name,email,phone&limit=5000",
        headers,
    )
    if status != 200 or not isinstance(existing, list):
        print(f"FAIL reading event_clients: status={status} body={str(existing)[:300]}")
        return 1

    have = {
        client_identity(
            r.get("company_name"),
            r.get("contact_name"),
            r.get("email"),
            normalize_phone(r.get("phone")),
        )
        for r in existing
    }

    now = datetime.now(timezone.utc).isoformat()
    to_insert = []
    for r in rows:
        ident = client_identity(r["company"], r.get("contact"), r.get("email"), r.get("phone"))
        if ident in have:
            continue
        to_insert.append(
            {
                "company_name": r["company"],
                "contact_name": r.get("contact"),
                "email": r.get("email"),
                "phone": r.get("phone"),
                "source": "excel_seed",
                "owner_username": "seed",
                "created_at": now,
                "updated_at": now,
            }
        )

    inserted = 0
    chunk = 80
    for i in range(0, len(to_insert), chunk):
        slice_ = to_insert[i : i + chunk]
        st, body = http_json(
            "POST",
            f"{url}/rest/v1/event_clients",
            headers,
            slice_,
        )
        if not (200 <= st < 300):
            print(f"FAIL insert at offset {i}: status={st} body={str(body)[:300]}")
            return 1
        inserted += len(slice_)

    print(
        f"Supabase OK inserted={inserted} skipped={len(rows) - inserted} "
        f"totalSeed={len(rows)} existingBefore={len(existing)}"
    )
    return 0


def main() -> int:
    parser = argparse.ArgumentParser(description="Excel → seed_event_clients.json")
    parser.add_argument("--excel", type=Path, default=DEFAULT_EXCEL)
    parser.add_argument("--out", type=Path, default=SEED)
    parser.add_argument(
        "--seed",
        action="store_true",
        help="También insertar faltantes en Supabase (service role)",
    )
    args = parser.parse_args()

    rows = read_excel(args.excel)
    args.out.parent.mkdir(parents=True, exist_ok=True)
    args.out.write_text(
        json.dumps(rows, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    print(f"OK wrote {args.out} ({len(rows)} clientes)")

    if args.seed:
        return seed_supabase(rows)

    print(
        "Tip: sube a Supabase con --seed, o POST /api/eventos/clients/seed "
        "(sesión con permiso de edición), o python scripts/seed_event_clients.py"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
