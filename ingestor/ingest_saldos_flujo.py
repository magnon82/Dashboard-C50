"""
Extrae del FLUJO EFECTIVO CARRANZA 50.xlsx → Supabase:
  1) flujo_efectivo_saldo  — saldo diario (columna Saldo en Efectivo)
  2) flujo_efectivo_semana — ingresos/egresos por semana de presupuesto
  3) flujo_efectivo_mov    — movimientos línea a línea (caja chica, otros, etc.)

La semana de cada movimiento se toma del texto en Concepto
(p. ej. \"EFECTIVO SEMANA #26\", \"CAJA CHICA SEM #7\"). Si Concepto no
indica semana, se usa la fecha. Los números de semana anuales se mapean
a SEM 1–5 del mes (mismo criterio que el resumen bancario).

Las filas CAJA CHICA SEMANA #N son el desglose semanal del presupuesto
en efectivo (los totales coinciden con la columna Efectivo del presupuesto).

CLI:
  python ingest_saldos_flujo.py              # todos los años
  python ingest_saldos_flujo.py --year 2026
  python ingest_saldos_flujo.py --dry-run
  python sync_saldos_al_dia.py              # wrapper (Actions cada 5 min)
"""

from __future__ import annotations

import argparse
import json
import os
import re
import tempfile
from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path

from dotenv import load_dotenv
from openpyxl import load_workbook
from supabase import create_client

load_dotenv()
load_dotenv(Path(__file__).resolve().parent.parent / ".env.local")

SOURCE_FILE = "flujo_efectivo_saldo"
SOURCE_SEMANA = "flujo_efectivo_semana"
SOURCE_MOV = "flujo_efectivo_mov"
DEFAULT_PATH = Path(r"I:\Mi unidad\Administración\FLUJO EFECTIVO CARRANZA 50.xlsx")
DRIVE_FILE_NAME = "FLUJO EFECTIVO CARRANZA 50.xlsx"
COL_FECHA = 2
COL_CONCEPTO = 3
COL_VENTAS = 4
COL_OTROS_ING = 5
COL_CAJA_CHICA = 6
COL_OTROS_EGR = 7
COL_FINIQUITOS = 8
COL_SALDO = 9

# (columna 1-based, type, category, col_key)
AMOUNT_COLUMNS = (
    (COL_VENTAS, "income", "Ventas", "ventas"),
    (COL_OTROS_ING, "income", "Otros ingresos", "otros_ingresos"),
    (COL_CAJA_CHICA, "expense", "Caja chica", "caja_chica"),
    (COL_OTROS_EGR, "expense", "Otros egresos", "otros_egresos"),
    (COL_FINIQUITOS, "expense", "Finiquitos / Uniformes", "finiquitos"),
)

RE_CAJA_CHICA = re.compile(r"CAJA\s*CHICA", re.IGNORECASE)

# Semana anual tipo Acumulado / Concepto: SEMANA #N, SEM #N, SEMANA N, SEM N
RE_ANNUAL_WEEK = re.compile(
    r"(?<![A-ZÁÉÍÓÚ])SEM(?:ANA)?\s*#?\s*(\d+)\b",
    re.IGNORECASE,
)
# Evitar "2 SEMANAS", "4 SEMANAS DE EVENTOS"
RE_FALSE_SEMANAS = re.compile(r"\d+\s+SEMANAS?\b", re.IGNORECASE)

MONTH_ALIASES: dict[str, int] = {
    "ENE": 1,
    "ENERO": 1,
    "FEB": 2,
    "FEBRERO": 2,
    "MAR": 3,
    "MARZO": 3,
    "ABR": 4,
    "ABRIL": 4,
    "MAY": 5,
    "MAYO": 5,
    "JUN": 6,
    "JUNIO": 6,
    "JUL": 7,
    "JULIO": 7,
    "AGO": 8,
    "AGOSTO": 8,
    "SEP": 9,
    "SEPT": 9,
    "SEPTIEMBRE": 9,
    "OCT": 10,
    "OCTUBRE": 10,
    "NOV": 11,
    "NOVIEMBRE": 11,
    "DIC": 12,
    "DICIEMBRE": 12,
}

# Mes + semana relativa del mes: "JULIO SEM 2", "ENE SEMANA #1"
RE_MONTH_WEEK = re.compile(
    r"\b("
    + "|".join(sorted(MONTH_ALIASES.keys(), key=len, reverse=True))
    + r")\b\s*SEM(?:ANA)?\s*#?\s*(\d+)\b",
    re.IGNORECASE,
)


def resolve_flujo_path(path: Path) -> Path:
    """Usa ruta local si existe; si no, descarga desde Google Drive."""
    if path.exists():
        return path
    from google_auth import download_drive_file_by_name

    print(f"No hay archivo local ({path}); descargando desde Drive…")
    dest = Path(tempfile.gettempdir()) / DRIVE_FILE_NAME
    downloaded = download_drive_file_by_name(DRIVE_FILE_NAME, dest)
    print(f"Descargado: {downloaded}")
    return downloaded


def sheet_name_for_year(year: int) -> str:
    if year == 2024:
        return "FUJO DE EFECTIVO 2024"
    return f"FLUJO DE EFECTIVO {year}"


def as_date(value) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def as_amount(value) -> float | None:
    if value is None or isinstance(value, str):
        return None
    try:
        amount = float(value)
    except (TypeError, ValueError):
        return None
    if amount == 0:
        return None
    return abs(amount)


def years_in_workbook(path: Path) -> list[int]:
    wb = load_workbook(path, read_only=True, data_only=True)
    years: list[int] = []
    for name in wb.sheetnames:
        m = re.search(r"(20\d{2})", name)
        if m:
            years.append(int(m.group(1)))
    wb.close()
    return sorted(set(years))


def valid_date_for_sheet(fecha: date, sheet_year: int) -> bool:
    """Acepta fechas del año de la hoja o enero del año siguiente (corte anual)."""
    if fecha.year == sheet_year:
        return True
    if fecha.year == sheet_year + 1 and fecha.month == 1:
        return True
    # Filas de cierre del año anterior pegadas al inicio de la hoja
    if fecha.year == sheet_year - 1 and fecha.month == 12:
        return True
    return False


def first_monday_on_or_after(year: int, month: int, day: int) -> date:
    d = date(year, month, day)
    while d.weekday() != 0:
        d += timedelta(days=1)
    return d


def first_monday_jan1(year: int) -> date:
    return first_monday_on_or_after(year, 1, 1)


def monday_of_annual_week(year: int, week: int) -> date:
    return first_monday_jan1(year) + timedelta(days=(week - 1) * 7)


def annual_week_for_date(fecha: date) -> int:
    """# de semana alineado a Acumulado / Concepto (primer lunes ≥ 1 ene)."""
    week1 = first_monday_jan1(fecha.year)
    # Lunes de la semana de fecha
    mon = fecha - timedelta(days=fecha.weekday())
    if mon < week1:
        # Días previos al lun 1 → semana del año anterior
        prev = first_monday_jan1(fecha.year - 1)
        return (mon - prev).days // 7 + 1
    return (mon - week1).days // 7 + 1


def resolve_year_for_annual_week(
    week_num: int, fecha: date, sheet_year: int
) -> int:
    """
    Elige el año del # de semana en Concepto.
    Útil para SEMANA #51–#53 registradas en enero del año siguiente.
    """
    candidates = {fecha.year, fecha.year - 1, sheet_year, sheet_year - 1}
    best_y = fecha.year
    best_score = None
    for y in candidates:
        if y < 2000 or week_num < 1:
            continue
        mon = monday_of_annual_week(y, week_num)
        dist = (fecha - mon).days
        # Lunes no más de 14 días después de la fecha; no más de ~90 días antes
        if dist < -14 or dist > 90:
            continue
        score = abs(dist)
        if best_score is None or score < best_score:
            best_score = score
            best_y = y
    return best_y


def month_sem_for_annual_week(year: int, week: int) -> tuple[int, int, int] | None:
    """Mapea semana anual → (año, mes, SEM n del mes presupuesto)."""
    if week < 1:
        return None
    mon = monday_of_annual_week(year, week)
    y, m = mon.year, mon.month
    mstart = first_monday_on_or_after(y, m, 1)
    if mon < mstart:
        if m == 1:
            y, m = y - 1, 12
        else:
            m -= 1
        mstart = first_monday_on_or_after(y, m, 1)
    idx = (mon - mstart).days // 7 + 1
    if idx < 1 or idx > 6:
        return None
    return y, m, idx


def parse_concepto_week(
    concepto: str, fecha: date, sheet_year: int
) -> tuple[int, int, int, str]:
    """
    Devuelve (year, month, week_sem, source) donde source es
    'concepto_mes' | 'concepto' | 'fecha'.
    """
    text = (concepto or "").strip()

    # Mes nombrado + semana relativa (raro en el archivo; best-effort)
    mm = RE_MONTH_WEEK.search(text)
    if mm:
        month = MONTH_ALIASES.get(mm.group(1).upper())
        w = int(mm.group(2))
        if month and 1 <= w <= 6:
            y = fecha.year
            if month == 12 and fecha.month <= 2:
                y = fecha.year - 1
            elif month <= 2 and fecha.month >= 11:
                y = fecha.year + 1
            return y, month, w, "concepto_mes"

    # Semana anual en Concepto (patrón dominante: SEMANA #N / SEM #N)
    m = RE_ANNUAL_WEEK.search(text)
    if m:
        # "2 SEMANAS" / "4 SEMANAS DE…" sin "#N" no cuentan
        false_plural = RE_FALSE_SEMANAS.search(text)
        has_hash = re.search(r"SEM(?:ANA)?\s*#\s*\d+", text, re.I)
        if not false_plural or has_hash:
            week_num = int(m.group(1))
            if week_num >= 1:
                y = resolve_year_for_annual_week(week_num, fecha, sheet_year)
                mapped = month_sem_for_annual_week(y, week_num)
                if mapped:
                    return (*mapped, "concepto")

    # Fallback: fecha → semana anual → SEM del mes
    week_num = annual_week_for_date(fecha)
    if week_num < 1:
        mon = fecha - timedelta(days=fecha.weekday())
        prev_year = fecha.year - 1
        week_num = (mon - first_monday_jan1(prev_year)).days // 7 + 1
        mapped = month_sem_for_annual_week(prev_year, week_num)
    else:
        mapped = month_sem_for_annual_week(fecha.year, week_num)
    if mapped:
        return mapped[0], mapped[1], mapped[2], "fecha"
    return fecha.year, fecha.month, 1, "fecha"


def extract_saldos(path: Path, year: int) -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    name = sheet_name_for_year(year)
    if name not in wb.sheetnames:
        wb.close()
        return []

    ws = wb[name]
    records: list[dict] = []
    for row in ws.iter_rows(min_row=3, max_col=COL_SALDO, values_only=True):
        cells = list(row) + [None] * COL_SALDO
        fecha = as_date(cells[COL_FECHA - 1])
        saldo = cells[COL_SALDO - 1]
        concepto = cells[COL_CONCEPTO - 1]
        if not fecha or saldo is None:
            continue
        if not valid_date_for_sheet(fecha, year):
            continue
        try:
            amount = float(saldo)
        except (TypeError, ValueError):
            continue
        records.append(
            {
                "date": fecha.isoformat(),
                "type": "income",
                "category": "Saldo Efectivo",
                "amount": amount,
                "description": f"FLUJO EFECTIVO CARRANZA 50 · {concepto or 'movimiento'}",
                "source_file": SOURCE_FILE,
            }
        )
    wb.close()
    return records


def extract_semana_efectivo(path: Path, year: int) -> list[dict]:
    """Agrega ingresos/egresos por (año, mes, SEM) según Concepto."""
    wb = load_workbook(path, read_only=True, data_only=True)
    name = sheet_name_for_year(year)
    if name not in wb.sheetnames:
        wb.close()
        return []

    ws = wb[name]
    # key -> {ingresos, egresos, from_concepto, from_fecha}
    buckets: dict[tuple[int, int, int], dict[str, float]] = defaultdict(
        lambda: {
            "ingresos": 0.0,
            "egresos": 0.0,
            "from_concepto": 0.0,
            "from_fecha": 0.0,
        }
    )

    for row in ws.iter_rows(min_row=3, max_col=COL_FINIQUITOS, values_only=True):
        cells = list(row) + [None] * COL_FINIQUITOS
        fecha = as_date(cells[COL_FECHA - 1])
        concepto_raw = cells[COL_CONCEPTO - 1]
        if not fecha or concepto_raw is None:
            continue
        concepto = str(concepto_raw).strip()
        if not concepto or concepto.upper().startswith("SALDO INICIAL"):
            continue
        if not valid_date_for_sheet(fecha, year):
            continue

        ingresos = 0.0
        egresos = 0.0
        for col in (COL_VENTAS, COL_OTROS_ING):
            amt = as_amount(cells[col - 1])
            if amt:
                ingresos += amt
        for col in (COL_CAJA_CHICA, COL_OTROS_EGR, COL_FINIQUITOS):
            amt = as_amount(cells[col - 1])
            if amt:
                egresos += amt
        if ingresos == 0 and egresos == 0:
            continue

        y, m, w, src = parse_concepto_week(concepto, fecha, year)
        key = (y, m, w)
        buckets[key]["ingresos"] += ingresos
        buckets[key]["egresos"] += egresos
        if src.startswith("concepto"):
            buckets[key]["from_concepto"] += 1
        else:
            buckets[key]["from_fecha"] += 1

    wb.close()

    records: list[dict] = []
    for (y, m, w), vals in sorted(buckets.items()):
        neto = vals["ingresos"] - vals["egresos"]
        payload = {
            "week": w,
            "year": y,
            "month": m,
            "efectivo_ingresos": round(vals["ingresos"], 2),
            "efectivo_egresos": round(vals["egresos"], 2),
            "efectivo_neto": round(neto, 2),
            "rows_from_concepto": int(vals["from_concepto"]),
            "rows_from_fecha": int(vals["from_fecha"]),
        }
        records.append(
            {
                "date": f"{y:04d}-{m:02d}-01",
                "type": "expense",
                "category": f"Efectivo Semana {w}",
                "amount": abs(neto),
                "description": json.dumps(payload, ensure_ascii=False),
                "source_file": SOURCE_SEMANA,
            }
        )
    return records


def annual_week_num_from_concepto(
    concepto: str, fecha: date, sheet_year: int
) -> int | None:
    """Extrae # de semana anual del Concepto si existe; si no, None."""
    text = (concepto or "").strip()
    m = RE_ANNUAL_WEEK.search(text)
    if not m:
        return None
    false_plural = RE_FALSE_SEMANAS.search(text)
    has_hash = re.search(r"SEM(?:ANA)?\s*#\s*\d+", text, re.I)
    if false_plural and not has_hash:
        return None
    week_num = int(m.group(1))
    if week_num < 1:
        return None
    return week_num


def extract_movimientos(path: Path, year: int) -> list[dict]:
    """
    Movimientos línea a línea del flujo de efectivo.
    Una fila Excel con montos en varias columnas → un registro por columna.
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    name = sheet_name_for_year(year)
    if name not in wb.sheetnames:
        wb.close()
        return []

    ws = wb[name]
    records: list[dict] = []

    for row in ws.iter_rows(min_row=3, max_col=COL_SALDO, values_only=True):
        cells = list(row) + [None] * COL_SALDO
        fecha = as_date(cells[COL_FECHA - 1])
        concepto_raw = cells[COL_CONCEPTO - 1]
        if not fecha or concepto_raw is None:
            continue
        concepto = str(concepto_raw).strip()
        if not concepto or concepto.upper().startswith("SALDO INICIAL"):
            continue
        if not valid_date_for_sheet(fecha, year):
            continue

        y, m, w, src = parse_concepto_week(concepto, fecha, year)
        week_annual = annual_week_num_from_concepto(concepto, fecha, year)
        if week_annual is None:
            week_annual = annual_week_for_date(fecha)

        saldo_raw = cells[COL_SALDO - 1]
        try:
            saldo_efectivo = float(saldo_raw) if saldo_raw is not None else None
        except (TypeError, ValueError):
            saldo_efectivo = None

        es_caja_concepto = bool(RE_CAJA_CHICA.search(concepto))

        for col_idx, tipo, categoria, col_key in AMOUNT_COLUMNS:
            amount = as_amount(cells[col_idx - 1])
            if amount is None:
                continue
            ingreso = amount if tipo == "income" else None
            egreso = amount if tipo == "expense" else None
            payload = {
                "canal": "EFECTIVO",
                "fecha": fecha.isoformat(),
                "concepto": concepto,
                "descripcion": concepto,
                "categoria": categoria,
                "columna": col_key,
                "ingreso": ingreso,
                "egreso": egreso,
                "cargo": egreso,
                "abono": ingreso,
                "week": w,
                "week_annual": week_annual,
                "year": y,
                "month": m,
                "week_source": src,
                "es_caja_chica": col_key == "caja_chica" or es_caja_concepto,
                "saldo_efectivo": saldo_efectivo,
                "source_path": DRIVE_FILE_NAME,
            }
            records.append(
                {
                    "date": fecha.isoformat(),
                    "type": tipo,
                    "category": categoria,
                    "amount": amount,
                    "description": json.dumps(payload, ensure_ascii=False),
                    "source_file": SOURCE_MOV,
                }
            )

    wb.close()
    return records


def extract_all(
    path: Path, years: list[int] | None
) -> tuple[list[dict], list[dict], list[dict]]:
    target = years or years_in_workbook(path)
    all_saldos: list[dict] = []
    all_semanas: list[dict] = []
    all_movs: list[dict] = []
    for year in target:
        rows = extract_saldos(path, year)
        weeks = extract_semana_efectivo(path, year)
        movs = extract_movimientos(path, year)
        if rows:
            print(f"  {year}: {len(rows)} saldos · último {rows[-1]['date']}")
        if weeks:
            print(f"  {year}: {len(weeks)} semanas efectivo")
        if movs:
            egresos = sum(1 for r in movs if r["type"] == "expense")
            print(f"  {year}: {len(movs)} movimientos línea ({egresos} egresos)")
        all_saldos.extend(rows)
        all_semanas.extend(weeks)
        all_movs.extend(movs)
    return all_saldos, all_semanas, all_movs


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Ingesta saldos + efectivo semanal + movimientos (FLUJO EFECTIVO) → Supabase"
    )
    parser.add_argument("--year", type=int, default=None, help="Un solo año (opcional)")
    parser.add_argument("--file", type=Path, default=DEFAULT_PATH)
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    path = resolve_flujo_path(args.file)
    years = [args.year] if args.year else None
    saldos, semanas, movs = extract_all(path, years)
    latest = max(saldos, key=lambda r: r["date"]) if saldos else None
    print(f"TOTAL registros saldo: {len(saldos)}")
    print(f"TOTAL registros semana efectivo: {len(semanas)}")
    print(f"TOTAL registros movimientos: {len(movs)}")
    if latest:
        print(f"Último día en archivo: {latest['date']} = ${latest['amount']:,.2f}")
    if semanas:
        print("Ejemplo semana:", semanas[-1])
    if movs:
        print("Ejemplo movimiento:", movs[-1])

    if args.dry_run:
        return

    url = os.environ.get("SUPABASE_URL") or os.environ.get("NEXT_PUBLIC_SUPABASE_URL")
    key = (
        os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
        or os.environ.get("SUPABASE_ANON_KEY")
        or os.environ.get("NEXT_PUBLIC_SUPABASE_ANON_KEY")
    )
    if not url or not key:
        raise SystemExit("Faltan variables en .env / .env.local")

    supabase = create_client(url, key)
    for source in (SOURCE_FILE, SOURCE_SEMANA, SOURCE_MOV):
        supabase.table("financial_records").delete().eq("source_file", source).execute()
    for batch_src, label in (
        (saldos, "saldos"),
        (semanas, "semanas"),
        (movs, "movimientos"),
    ):
        for i in range(0, len(batch_src), 200):
            supabase.table("financial_records").insert(batch_src[i : i + 200]).execute()
        print(f"Insertados {label}: {len(batch_src)}")


if __name__ == "__main__":
    main()
