"""
Ingestor Bloque B: lee FLUJO EFECTIVO CARRANZA 50.xlsx
y sube entradas/salidas a financial_records en Supabase.
"""

from __future__ import annotations

import argparse
import os
from datetime import date, datetime
from pathlib import Path

from dotenv import load_dotenv
from openpyxl import load_workbook
from supabase import create_client

load_dotenv()

DEFAULT_PATH = Path(r"I:\Mi unidad\Administración\FLUJO EFECTIVO CARRANZA 50.xlsx")
SOURCE_FILE = "flujo_efectivo"

# Columnas (1-based) según encabezado de la hoja
COL_FECHA = 2
COL_CONCEPTO = 3
# Ingresos
COL_VENTAS = 4
COL_OTROS_ING = 5
# Egresos
COL_CAJA_CHICA = 6
COL_OTROS_EGR = 7
COL_FINIQUITOS = 8

AMOUNT_COLUMNS = (
    (COL_VENTAS, "income", "Ventas"),
    (COL_OTROS_ING, "income", "Otros ingresos"),
    (COL_CAJA_CHICA, "expense", "Caja chica"),
    (COL_OTROS_EGR, "expense", "Otros egresos"),
    (COL_FINIQUITOS, "expense", "Finiquitos / Uniformes"),
)


def sheet_name_for_year(year: int) -> str:
    # 2024 tiene typo en el archivo: "FUJO DE EFECTIVO 2024"
    if year == 2024:
        return "FUJO DE EFECTIVO 2024"
    return f"FLUJO DE EFECTIVO {year}"


def as_date(value) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def as_amount(value) -> float | None:
    if value is None or isinstance(value, str):
        return None
    try:
        amount = float(value)
    except (TypeError, ValueError):
        return None
    if amount == 0:
        return None
    return abs(amount)


def extract_rows(workbook_path: Path, year: int) -> list[dict]:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    name = sheet_name_for_year(year)
    if name not in wb.sheetnames:
        wb.close()
        raise SystemExit(f"No existe la hoja '{name}'. Hojas: {wb.sheetnames}")

    ws = wb[name]
    records: list[dict] = []

    for row in ws.iter_rows(min_row=3, max_col=8, values_only=True):
        fecha = as_date(row[COL_FECHA - 1])
        concepto = row[COL_CONCEPTO - 1]
        if not fecha or not concepto or not str(concepto).strip():
            continue

        concepto_txt = str(concepto).strip()
        # Fila de etiquetas / saldo inicial sin montos reales en columnas de movimiento
        if concepto_txt.upper().startswith("SALDO INICIAL"):
            continue

        for col_idx, tipo, categoria in AMOUNT_COLUMNS:
            amount = as_amount(row[col_idx - 1])
            if amount is None:
                continue
            records.append(
                {
                    "date": fecha.isoformat(),
                    "type": tipo,
                    "category": categoria,
                    "amount": amount,
                    "description": concepto_txt,
                    "source_file": SOURCE_FILE,
                }
            )

    wb.close()
    return records


def chunked(items: list[dict], size: int = 200):
    for i in range(0, len(items), size):
        yield items[i : i + size]


def main() -> None:
    parser = argparse.ArgumentParser(description="Ingesta flujo de efectivo → Supabase")
    parser.add_argument("--year", type=int, default=2026, help="Año de la hoja a ingerir")
    parser.add_argument(
        "--file",
        type=Path,
        default=DEFAULT_PATH,
        help="Ruta al Excel de flujo de efectivo",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Solo muestra conteos, no escribe en Supabase",
    )
    args = parser.parse_args()

    if not args.file.exists():
        raise SystemExit(f"No se encontró el archivo: {args.file}")

    url = os.environ.get("SUPABASE_URL")
    key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
    if not url or not key:
        raise SystemExit("Faltan SUPABASE_URL o SUPABASE_SERVICE_ROLE_KEY en .env")

    records = extract_rows(args.file, args.year)
    incomes = sum(1 for r in records if r["type"] == "income")
    expenses = sum(1 for r in records if r["type"] == "expense")
    total_in = sum(r["amount"] for r in records if r["type"] == "income")
    total_out = sum(r["amount"] for r in records if r["type"] == "expense")

    print(f"Archivo: {args.file}")
    print(f"Hoja: {sheet_name_for_year(args.year)}")
    print(f"Registros: {len(records)} (ingresos={incomes}, egresos={expenses})")
    print(f"Totales: ingresos=${total_in:,.2f} | egresos=${total_out:,.2f}")

    if args.dry_run:
        print("Dry-run: no se escribió nada.")
        if records:
            print("Ejemplo:", records[0])
        return

    supabase = create_client(url, key)

    # Reemplazo idempotente de esta fuente (solo flujo_efectivo)
    deleted = (
        supabase.table("financial_records")
        .delete()
        .eq("source_file", SOURCE_FILE)
        .execute()
    )
    print(f"Limpieza previa source_file={SOURCE_FILE}: OK")

    inserted = 0
    for batch in chunked(records, 200):
        result = supabase.table("financial_records").insert(batch).execute()
        inserted += len(result.data or [])

    print(f"Insertados: {inserted}")
    print("Listo. Recarga http://localhost:3000")


if __name__ == "__main__":
    main()
