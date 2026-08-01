"""
Ingestor: estados de cuenta anuales MIFEL / BBVA → financial_records.

Fuente (Drive local):
  I:\\Mi unidad\\COMPROBANTES BANCARIOS\\{year}\\Estado de cuenta MIFEL {year}.xlsx
  I:\\Mi unidad\\COMPROBANTES BANCARIOS\\{year}\\Estado de cuenta BBVA {year}.xlsx

Columnas esperadas (fila de encabezado):
  Fecha, Descripción, Folio, Referencia, Cargo, Abono, Saldo total, RFC, IVA, Cheque

source_file:
  - estado_mifel
  - estado_bbva

description (JSON): columnas + bank + matched_rubro / match_confidence / observaciones / match_status

Uso:
  cd ingestor
  python ingest_estados_cuenta.py
  python ingest_estados_cuenta.py --year 2026 --dry-run
  python ingest_estados_cuenta.py --bank mifel --file "ruta.xlsx"
  python ingest_estados_cuenta.py --index-pdfs   # índice ligero PDFs de pagos (COMPROBANTES BANCARIOS)
  python ingest_estados_cuenta.py --index-pdfs --pdf-only --pdf-years 2023,2024,2025,2026
  python ingest_estados_cuenta.py --index-estados-pdf --pdf-only  # PDFs estados en Administración\\Bancos
"""

from __future__ import annotations

import argparse
import difflib
import json
import os
import re
import unicodedata
from datetime import date, datetime, timedelta
from pathlib import Path

from dotenv import load_dotenv
from openpyxl import load_workbook
from supabase import create_client

load_dotenv()
load_dotenv(Path(__file__).resolve().parent.parent / ".env.local")

SOURCE_MIFEL = "estado_mifel"
SOURCE_BBVA = "estado_bbva"
SOURCE_PDF_INDEX = "estado_pdf_index"
SOURCE_ESTADO_CUENTA_PDF_INDEX = "estado_cuenta_pdf_index"

DEFAULT_COMPROBANTES = Path(r"I:\Mi unidad\COMPROBANTES BANCARIOS")
DEFAULT_BANCOS = Path(r"I:\Mi unidad\Administración\Bancos")
DEFAULT_ESTADOS_PDF = DEFAULT_BANCOS / "Mifel" / "Estados de cuenta"

HEADER_ALIASES = {
    "fecha": "fecha",
    "descripcion": "descripcion",
    "descripción": "descripcion",
    "folio": "folio",
    "referencia": "referencia",
    "cargo": "cargo",
    "cargo (mxn)": "cargo",
    "abono": "abono",
    "abono (mxn)": "abono",
    "saldo total": "saldo_total",
    "saldo total (mxn)": "saldo_total",
    "rfc": "rfc",
    "iva": "iva",
    "iva(mxn)": "iva",
    "iva (mxn)": "iva",
    "cheque": "cheque",
}

# (keywords in description, rubro, parent|None, confidence)
VENDOR_RULES: list[tuple[tuple[str, ...], str, str | None, float]] = [
    (("frutas y verduras", "frutasyverduras", "provee", "del campo"), "Frutas y verduras", "Insumos de cocina", 0.9),
    (("carnes premium", "proteinas", "carnes"), "Proteínas", "Insumos de cocina", 0.85),
    (("lacteos", "lácteos", "osmi", "mante"), "Lácteos", "Insumos de cocina", 0.8),
    (("panes", "tortillas", "postres", "molino"), "Panes, tortillas, postres", "Insumos de cocina", 0.8),
    (("carbon", "carbón"), "Carbón", "Insumos de cocina", 0.9),
    (("hielo", "fabrica de hielo", "aguas y hielo"), "Refrescos, aguas y hielo", "Insumos de barra", 0.75),
    (("heineken", "cervezas cuauhtemoc", "cerveza", "cerveceria", "la bru"), "Cervezas", "Insumos de barra", 0.9),
    (("destilados", "vinos", "castellana", "vina de santiago", "viña"), "Destilados y vinos", "Insumos de barra", 0.85),
    (("penafiel", "peñafiel", "refrescos", "hercules", "craft galaxi"), "Refrescos, aguas y hielo", "Insumos de barra", 0.8),
    (("cafe", "café", "tutuka"), "Café", "Insumos de barra", 0.75),
    (("lavanderia", "lavandería"), "Lavandería", "Servicios", 0.95),
    (("telmex", "telefonos de mexico", "teléfono", "telefono"), "Teléfono", "Servicios", 0.95),
    (("cea", "comision estatal de aguas", "agua "), "Agua", "Servicios", 0.85),
    (("engie", "gas ", "tractebel"), "Gas", "Servicios", 0.9),
    (("cfe", "luz ", "energia electrica"), "Luz", "Servicios", 0.9),
    (("contador", "joseantonio mendoza", "mendoza pesquera"), "Contador", "Servicios", 0.9),
    (("alarma", "soluciones de alta"), "Alarma", "Servicios", 0.9),
    (("diseño", "diseno", "publicidad", "redes", "lorena rodriguez"), "Diseño y publicidad", "Servicios", 0.8),
    (("auditoria",), "Auditorías", "Servicios", 0.85),
    (("materias primas",), "Materias primas", "Servicios", 0.85),
    (("imss",), "IMSS", None, 0.95),
    (
        (
            "impuestos",
            "sat",
            "isr",
            "iva",
            "hacienda",
            "shcp",
            "tesoreria",
            "tesorería",
            "secretaria de hacienda",
            "infonavit",
            "linea de captura",
            "línea de captura",
        ),
        "Impuestos",
        None,
        0.85,
    ),
    (("infonavit",), "Impuestos", None, 0.9),
    (("nomina", "nómina", "quincena", "administraciones zen-nom", "nomina meseros"), "Nómina", None, 0.9),
    (("renta",), "Renta", None, 0.95),
    (("cristaleria", "cristalería", "equipo"), "Cristalería y Equipo", None, 0.9),
    (("mantenimiento",), "Mantenimiento", None, 0.85),
    (("papeleria", "papelería"), "Papelería", None, 0.85),
    (("limpieza", "baños", "biolimpieza", "technoclean", "sanitizante"), "Limpieza y baños", None, 0.85),
    (("gasolina", "taxi", "uber"), "Gasolina y taxis", None, 0.8),
    (("comida personal", "reembolso comida"), "Comida personal", None, 0.8),
    (("comision", "comisión bancaria"), "Comisiones bancarias", None, 0.8),
    (("finiquito", "reclutamiento"), "Finiquitos y reclutamiento", None, 0.85),
    (("licencia", "afiliacion", "afiliación"), "Licencias y afiliaciones", None, 0.8),
]

RUBRO_NAMES = [
    "Insumos de cocina",
    "Frutas y verduras",
    "Proteínas",
    "Abarrotes",
    "Lácteos",
    "Panes, tortillas, postres",
    "Agua",
    "Carbón",
    "Insumos de barra",
    "Destilados y vinos",
    "Cervezas",
    "Café",
    "Refrescos, aguas y hielo",
    "Servicios",
    "Lavandería",
    "Gas",
    "Luz",
    "Teléfono",
    "Contador",
    "Diseño y publicidad",
    "Alarma",
    "Auditorías",
    "Gas calentadores",
    "Materias primas",
    "Comida personal",
    "Renta",
    "Mantenimiento",
    "Cristalería y Equipo",
    "Papelería",
    "Limpieza y baños",
    "Gasolina y taxis",
    "Otros",
    "Licencias y afiliaciones",
    "Comisiones bancarias",
    "Finiquitos y reclutamiento",
    "Nómina",
    "IMSS",
    "Impuestos",
]


def norm_key(name: str) -> str:
    nfkd = unicodedata.normalize("NFD", name or "")
    plain = "".join(c for c in nfkd if unicodedata.category(c) != "Mn")
    return re.sub(r"[^A-Z0-9]+", " ", plain.upper()).strip()


def norm_compact(name: str) -> str:
    return norm_key(name).replace(" ", "")


def header_key(cell) -> str | None:
    if cell is None:
        return None
    raw = str(cell).strip().lower()
    raw = re.sub(r"\s+", " ", raw)
    return HEADER_ALIASES.get(raw)


def as_amount(value) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip().replace(",", "").replace("$", "").replace(" ", "")
    if not s or s == "-":
        return None
    try:
        return float(s)
    except ValueError:
        return None


def as_date(value) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s[:10], fmt).date()
        except ValueError:
            continue
    # Excel serial
    if isinstance(value, (int, float)):
        try:
            return (datetime(1899, 12, 30) + timedelta(days=float(value))).date()
        except (OverflowError, ValueError):
            return None
    return None


def match_rubro(descripcion: str) -> tuple[str | None, str | None, float]:
    text = norm_key(descripcion)
    compact = norm_compact(descripcion)
    if not text:
        return None, None, 0.0

    best: tuple[str | None, str | None, float] = (None, None, 0.0)

    for keywords, rubro, parent, conf in VENDOR_RULES:
        for kw in keywords:
            nk = norm_key(kw)
            nc = norm_compact(kw)
            if (nk and nk in text) or (nc and nc in compact):
                if conf > best[2]:
                    best = (rubro, parent, conf)

    # Direct rubro-name containment
    for rubro in RUBRO_NAMES:
        nk = norm_key(rubro)
        if len(nk) < 3:
            continue
        if nk in text or norm_compact(rubro) in compact:
            conf = 0.7 if len(nk) < 6 else 0.8
            if conf > best[2]:
                parent = None
                if nk in {
                    "FRUTAS Y VERDURAS",
                    "PROTEINAS",
                    "ABARROTES",
                    "LACTEOS",
                    "PANES TORTILLAS POSTRES",
                    "CARBON",
                }:
                    parent = "Insumos de cocina"
                elif nk in {
                    "DESTILADOS Y VINOS",
                    "CERVEZAS",
                    "CAFE",
                    "REFRESCOS AGUAS Y HIELO",
                }:
                    parent = "Insumos de barra"
                elif nk in {
                    "LAVANDERIA",
                    "AGUA",
                    "GAS",
                    "LUZ",
                    "TELEFONO",
                    "CONTADOR",
                    "DISENO Y PUBLICIDAD",
                    "ALARMA",
                    "AUDITORIAS",
                    "GAS CALENTADORES",
                    "MATERIAS PRIMAS",
                }:
                    parent = "Servicios"
                # Agua can be cocina or servicios — keep vendor rule preference
                if rubro == "Agua" and best[0] is None:
                    parent = "Servicios"
                best = (rubro, parent, conf)

    return best


def find_header_row(rows: list[list]) -> tuple[int, dict[str, int]] | None:
    for i, row in enumerate(rows[:40]):
        mapping: dict[str, int] = {}
        for col, cell in enumerate(row):
            key = header_key(cell)
            if key and key not in mapping:
                mapping[key] = col
        if "fecha" in mapping and ("descripcion" in mapping or "cargo" in mapping):
            return i, mapping
    return None


def cell_str(value) -> str | None:
    if value is None or value == "":
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip() or None


def parse_workbook(path: Path, bank: str) -> list[dict]:
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = [[cell.value for cell in row] for row in ws.iter_rows()]
    found = find_header_row(rows)
    if not found:
        print(f"  AVISO: no se encontró encabezado en {path.name}")
        return []
    header_idx, cols = found
    source = SOURCE_MIFEL if bank == "MIFEL" else SOURCE_BBVA
    records: list[dict] = []

    for row in rows[header_idx + 1 :]:
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue

        def get(key: str):
            idx = cols.get(key)
            if idx is None or idx >= len(row):
                return None
            return row[idx]

        fecha = as_date(get("fecha"))
        if not fecha:
            continue

        descripcion = cell_str(get("descripcion")) or ""
        cargo = as_amount(get("cargo"))
        abono = as_amount(get("abono"))
        # Prefer cargos as expenses; abonos as income (amount positive)
        if cargo and cargo != 0:
            amount = abs(cargo)
            rec_type = "expense"
        elif abono and abono != 0:
            amount = abs(abono)
            rec_type = "income"
        else:
            amount = 0.0
            rec_type = "expense"

        matched_rubro, matched_parent, confidence = match_rubro(descripcion)
        status = "matched" if matched_rubro and confidence >= 0.7 else "unmatched"

        payload = {
            "bank": bank,
            "fecha": fecha.isoformat(),
            "descripcion": descripcion,
            "folio": cell_str(get("folio")),
            "referencia": cell_str(get("referencia")),
            "cargo": cargo,
            "abono": abono,
            "saldo_total": as_amount(get("saldo_total")),
            "rfc": cell_str(get("rfc")),
            "iva": as_amount(get("iva")),
            "cheque": cell_str(get("cheque")),
            "matched_rubro": matched_rubro,
            "matched_parent": matched_parent,
            "match_confidence": round(confidence, 3),
            "match_status": status,
            "match_source": "auto",
            "observaciones": "",
            "source_path": path.name,
        }

        records.append(
            {
                "date": fecha.isoformat(),
                "type": rec_type,
                "category": matched_rubro or bank,
                "amount": amount,
                "description": json.dumps(payload, ensure_ascii=False),
                "source_file": source,
            }
        )

    return records


def resolve_bank_file(folder: Path, year: int, bank: str) -> Path | None:
    name = f"Estado de cuenta {bank} {year}.xlsx"
    path = folder / str(year) / name
    if path.exists():
        return path
    # Fuzzy search
    year_dir = folder / str(year)
    if not year_dir.is_dir():
        return None
    needle = bank.upper()
    for p in year_dir.glob("*.xlsx"):
        u = p.name.upper()
        if "ESTADO" in u and needle in u and str(year) in u:
            return p
    return None


# Amount may be prefixed with $ (usual) or bare digits.
PDF_NAME_RE = re.compile(
    r"^(?P<bank>mifel|bbva)[-_](?P<body>.+?)[-_]?\$?(?P<amount>[\d.,]+)\.pdf$",
    re.IGNORECASE,
)

# IMSS / impuestos / instituciones de gobierno (filenames + matching).
GOV_CONCEPTO_RE = re.compile(
    r"(imss|infonavit|shcp|hacienda|impuesto|tesorer|secretaria|sat\b|isr|iva|"
    r"linea\s*de\s*captura|l[ií]nea\s*de\s*captura)",
    re.IGNORECASE,
)


def is_gobierno_text(*parts: str) -> bool:
    blob = " ".join(p for p in parts if p)
    return bool(GOV_CONCEPTO_RE.search(blob))


def concepto_from_body(body: str) -> str:
    """Middle filename segment without bank/amount → readable Concepto.

    Examples:
      NominaMeserosSem28(26)  → Nomina Meseros Sem 28
      ProveedorXYZ-Sem 3      → Proveedor XYZ Sem 3
      Imss-Ene26              → IMSS Ene 26
      ImpuestosSAT-IsrEnero26 → Impuestos SAT ISR Enero 26
    """
    s = (body or "").strip()
    if not s:
        return ""
    # Drop year hints like (26) / (2026)
    s = re.sub(r"\(\d{2,4}\)", "", s)
    s = re.sub(r"[-_]+", " ", s)
    # Normalize SemN / Sem N
    s = re.sub(r"(?i)\bsem\s*(\d+)\b", r"Sem \1", s)
    # CamelCase / digit boundaries
    s = re.sub(r"(?<=[a-záéíóúñ])(?=[A-ZÁÉÍÓÚÑ])", " ", s)
    s = re.sub(r"(?<=[A-Za-zÁÉÍÓÚáéíóúñÑ])(?=\d)", " ", s)
    s = re.sub(r"(?<=\d)(?=[A-Za-zÁÉÍÓÚáéíóúñÑ])", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    # Canonical government / tax labels for search + UI
    replacements = (
        (r"(?i)\bimss\b", "IMSS"),
        (r"(?i)\binfonavit\b", "INFONAVIT"),
        (r"(?i)\bshcp\b", "SHCP"),
        (r"(?i)\bsat\b", "SAT"),
        (r"(?i)\bisr\b", "ISR"),
        (r"(?i)\biva\b", "IVA"),
        (r"(?i)\bimpuestos?\b", "Impuestos"),
        (r"(?i)\btesorer[ií]a\b", "Tesorería"),
        (r"(?i)\bsecretar[ií]a\s+de\s+hacienda\b", "Secretaría de Hacienda"),
    )
    for pat, repl in replacements:
        s = re.sub(pat, repl, s)
    return s


def _month_year_from_parent(parent: str, year_hint: int | None) -> tuple[int | None, int | None]:
    month = None
    year = year_hint
    upper = parent.upper()
    months = {
        "ENERO": 1,
        "FEBRERO": 2,
        "MARZO": 3,
        "ABRIL": 4,
        "MAYO": 5,
        "JUNIO": 6,
        "JULIO": 7,
        "AGOSTO": 8,
        "SEPTIEMBRE": 9,
        "OCTUBRE": 10,
        "NOVIEMBRE": 11,
        "DICIEMBRE": 12,
    }
    for name, num in months.items():
        if name in upper:
            month = num
            break
    ym = re.search(r"(20\d{2})", upper)
    if ym:
        year = int(ym.group(1))
    return month, year


def parse_pdf_filename(path: Path, year_hint: int | None = None) -> dict | None:
    """Index payment PDFs under COMPROBANTES BANCARIOS.

    Standard: MIFEL|BBVA-<concepto>-$<monto>.pdf
    Fallback: still index IMSS/impuestos/gobierno PDFs with odd names.
    """
    month, year = _month_year_from_parent(path.parent.name, year_hint)
    m = PDF_NAME_RE.match(path.name)
    if m:
        bank = m.group("bank").upper()
        body = m.group("body")
        amount_raw = m.group("amount").replace(",", "")
        try:
            amount = float(amount_raw)
        except ValueError:
            amount = 0.0
    elif is_gobierno_text(path.name):
        # Odd government receipt names — keep searchable in the index
        bank = (
            "BBVA"
            if "bbva" in path.name.lower()
            else "MIFEL"
            if "mifel" in path.name.lower()
            else ""
        )
        body = path.stem
        amount_m = re.search(r"\$?\s*([\d.,]+)\s*$", path.stem)
        try:
            amount = float(amount_m.group(1).replace(",", "")) if amount_m else 0.0
        except ValueError:
            amount = 0.0
    else:
        return None

    week_m = re.search(r"Sem\s*(\d+)", body, re.IGNORECASE)
    week = int(week_m.group(1)) if week_m else None

    vendor = body.split("-")[0].strip() if body else ""
    concepto = concepto_from_body(body)
    matched_rubro, matched_parent, confidence = match_rubro(body.replace("-", " "))
    if not matched_rubro and is_gobierno_text(body, path.name):
        # Prefer IMSS when IMSS appears; otherwise Impuestos for gov institutions
        if re.search(r"(?i)\bimss\b", f"{body} {path.name}"):
            matched_rubro, matched_parent, confidence = "IMSS", None, 0.9
        else:
            matched_rubro, matched_parent, confidence = "Impuestos", None, 0.85
    iso = f"{year or 2026:04d}-{month or 1:02d}-01"

    payload = {
        "bank": bank,
        "filename": path.name,
        "rel_path": str(path),
        "vendor": vendor,
        "body": body,
        "concepto": concepto,
        "week": week,
        "amount": amount,
        "matched_rubro": matched_rubro,
        "matched_parent": matched_parent,
        "match_confidence": round(confidence, 3),
        "match_status": "matched" if matched_rubro and confidence >= 0.7 else "unmatched",
        "index_only": True,
        "gobierno": bool(is_gobierno_text(body, path.name, concepto)),
    }
    return {
        "date": iso,
        "type": "expense",
        "category": matched_rubro or vendor or bank,
        "amount": amount,
        "description": json.dumps(payload, ensure_ascii=False),
        "source_file": SOURCE_PDF_INDEX,
    }


def index_pdfs(folder: Path, years: list[int] | None = None) -> list[dict]:
    records: list[dict] = []
    year_dirs = []
    if years:
        year_dirs = [folder / str(y) for y in years if (folder / str(y)).is_dir()]
    else:
        year_dirs = sorted(
            [p for p in folder.iterdir() if p.is_dir() and re.fullmatch(r"20\d{2}", p.name)],
            key=lambda p: p.name,
        )
    for yd in year_dirs:
        year = int(yd.name)
        for pdf in yd.rglob("*.pdf"):
            rec = parse_pdf_filename(pdf, year)
            if rec:
                records.append(rec)
    return records


_MONTHS_ES = {
    "ENERO": 1,
    "FEBRERO": 2,
    "MARZO": 3,
    "ABRIL": 4,
    "MAYO": 5,
    "JUNIO": 6,
    "JULIO": 7,
    "AGOSTO": 8,
    "SEPTIEMBRE": 9,
    "OCTUBRE": 10,
    "NOVIEMBRE": 11,
    "DICIEMBRE": 12,
}

# Common typos / variants seen in Drive filenames (e.g. JUNIIO, FEBRRERO).
_MONTH_ALIASES: dict[str, int] = {
    **_MONTHS_ES,
    "FEBRRERO": 2,
    "FEBREROO": 2,
    "MARSO": 3,
    "ABRILL": 4,
    "JUNIIO": 6,
    "JUNIOO": 6,
    "JUINIO": 6,
    "JULIOO": 7,
    "AGOSTOO": 8,
    "SETIEMBRE": 9,
    "SEPTEMBRE": 9,
    "OCTUBREE": 10,
}

_MONTH_CANON = list(_MONTHS_ES.keys())
_MONTH_SKIP_TOKENS = {
    "BBVA",
    "MIFEL",
    "AMEX",
    "AMERICAN",
    "EXPRESS",
    "ESTADOS",
    "CUENTA",
    "ESTADO",
    "BANCOS",
    "C50",
    "PDF",
}

_SKIP_ESTADO_PDF = re.compile(
    r"(gu[ií]a|operativa|administraci[oó]n corporativa|pagar[eé]|carta |solicitud|ine )",
    re.IGNORECASE,
)


def _detect_bank_from_text(*parts: str) -> str:
    u = " ".join(parts).upper()
    if "AMEX" in u or "AMERICAN EXPRESS" in u:
        return "AMEX"
    if "BBVA" in u:
        return "BBVA"
    if "MIFEL" in u:
        return "MIFEL"
    return ""


def _month_from_token(token: str) -> int | None:
    """Exact alias, then fuzzy match against Spanish month names (typos)."""
    nfkd = unicodedata.normalize("NFD", token.upper())
    plain = "".join(c for c in nfkd if unicodedata.category(c) != "Mn")
    plain = re.sub(r"[^A-Z]", "", plain)
    if not plain or plain in _MONTH_SKIP_TOKENS or len(plain) < 3:
        return None
    if plain in _MONTH_ALIASES:
        return _MONTH_ALIASES[plain]
    matches = difflib.get_close_matches(plain, _MONTH_CANON, n=1, cutoff=0.72)
    if matches:
        return _MONTHS_ES[matches[0]]
    return None


def _month_year_from_text(text: str) -> tuple[int | None, int | None]:
    """Extract month+year from a filename or folder segment.

    Prefers explicit month labels (incl. typos like JUNIIO) over nothing;
    does not invent January when the month is unknown.
    """
    upper = text.upper()
    month = None
    # Longest alias first so FEBRERO wins over FEB-like substrings if added later.
    for name, num in sorted(_MONTH_ALIASES.items(), key=lambda x: (-len(x[0]), x[0])):
        if name in upper:
            month = num
            break
    if month is None:
        for tok in re.findall(r"[A-ZÁÉÍÓÚÜÑ]{3,}", upper):
            month = _month_from_token(tok)
            if month is not None:
                break
    ym = re.search(r"(20\d{2})", upper)
    year = int(ym.group(1)) if ym else None
    return month, year


def parse_estado_cuenta_pdf(path: Path, year_hint: int | None = None) -> dict | None:
    """Parse monthly bank-statement PDFs under Administración\\Bancos."""
    if _SKIP_ESTADO_PDF.search(path.name):
        return None
    # Filename month/year wins over folder / mtime hints when present.
    name_m, name_y = _month_year_from_text(path.name)
    parent_m, parent_y = _month_year_from_text(path.parent.name)
    grand_m, grand_y = _month_year_from_text(path.parent.parent.name)
    year = name_y or parent_y or grand_y or year_hint
    month = name_m if name_m is not None else (parent_m or grand_m)
    bank = _detect_bank_from_text(path.name, path.parent.name, str(path))
    if not year:
        return None
    # Only default day; never invent January when month is unknown.
    if month is None:
        iso = f"{year:04d}-01-01"
    else:
        iso = f"{year:04d}-{month:02d}-01"
    payload = {
        "bank": bank,
        "filename": path.name,
        "rel_path": str(path),
        "month": month,
        "year": year,
        "index_only": True,
        "kind": "estado_cuenta",
    }
    return {
        "date": iso,
        "type": "expense",
        "category": bank or "estado_cuenta",
        "amount": 0,
        "description": json.dumps(payload, ensure_ascii=False),
        "source_file": SOURCE_ESTADO_CUENTA_PDF_INDEX,
    }


def index_estados_cuenta_pdfs(
    folder: Path, years: list[int] | None = None
) -> list[dict]:
    """Index PDFs under Mifel\\Estados de cuenta\\{year}\\..."""
    records: list[dict] = []
    if not folder.is_dir():
        print(f"AVISO: no existe carpeta de estados PDF: {folder}")
        return records

    year_dirs: list[Path] = []
    if years:
        year_dirs = [folder / str(y) for y in years if (folder / str(y)).is_dir()]
    else:
        year_dirs = sorted(
            [p for p in folder.iterdir() if p.is_dir() and re.fullmatch(r"20\d{2}", p.name)],
            key=lambda p: p.name,
        )

    for yd in year_dirs:
        year = int(yd.name)
        for pdf in yd.rglob("*.pdf"):
            rec = parse_estado_cuenta_pdf(pdf, year)
            if rec:
                records.append(rec)
    return records


def chunked(items: list, size: int = 200):
    for i in range(0, len(items), size):
        yield items[i : i + size]


def get_supabase():
    url = (
        os.environ.get("NEXT_PUBLIC_SUPABASE_URL")
        or os.environ.get("SUPABASE_URL")
        or ""
    ).strip()
    key = (
        os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
        or os.environ.get("SUPABASE_ANON_KEY")
        or os.environ.get("NEXT_PUBLIC_SUPABASE_ANON_KEY")
        or ""
    ).strip()
    if not url or not key:
        raise SystemExit("Faltan credenciales Supabase en .env / .env.local")
    return create_client(url, key)


def main() -> None:
    parser = argparse.ArgumentParser(description="Ingest estados de cuenta MIFEL/BBVA")
    parser.add_argument("--folder", type=Path, default=DEFAULT_COMPROBANTES)
    parser.add_argument("--year", type=int, default=date.today().year)
    parser.add_argument("--bank", choices=["mifel", "bbva", "both"], default="both")
    parser.add_argument("--file", type=Path, help="Excel puntual (requiere --bank mifel|bbva)")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument(
        "--index-pdfs",
        action="store_true",
        help="Índice ligero de PDFs de pagos (COMPROBANTES BANCARIOS) por nombre",
    )
    parser.add_argument(
        "--index-estados-pdf",
        action="store_true",
        help="Índice de PDFs de estados de cuenta en Administración\\Bancos",
    )
    parser.add_argument(
        "--estados-folder",
        type=Path,
        default=DEFAULT_ESTADOS_PDF,
        help="Carpeta raíz de estados PDF (default: ...\\Bancos\\Mifel\\Estados de cuenta)",
    )
    parser.add_argument(
        "--pdf-years",
        type=str,
        default="",
        help="Años CSV para índices PDF (ej. 2022,2023,2024,2025,2026). Vacío = todos.",
    )
    parser.add_argument(
        "--pdf-only",
        action="store_true",
        help="Solo índice PDF; no lee ni reemplaza estados Excel MIFEL/BBVA",
    )
    args = parser.parse_args()

    banks: list[str] = []
    if not args.pdf_only:
        if args.file:
            if args.bank == "both":
                raise SystemExit("Con --file indica --bank mifel o bbva")
            banks = [args.bank.upper()]
        elif args.bank == "both":
            banks = ["MIFEL", "BBVA"]
        else:
            banks = [args.bank.upper()]

    all_records: list[dict] = []
    sources_touched: set[str] = set()

    for bank in banks:
        if args.file:
            path = args.file
            if not path.exists():
                raise SystemExit(f"No existe: {path}")
        else:
            path = resolve_bank_file(args.folder, args.year, bank)
            if not path:
                print(f"AVISO: falta Estado de cuenta {bank} {args.year}.xlsx en {args.folder / str(args.year)}")
                continue
        print(f"Leyendo {bank}: {path}")
        rows = parse_workbook(path, bank)
        print(f"  Movimientos: {len(rows)}")
        matched = sum(
            1
            for r in rows
            if json.loads(r["description"]).get("match_status") == "matched"
        )
        print(f"  Auto-match: {matched}/{len(rows)}")
        all_records.extend(rows)
        sources_touched.add(SOURCE_MIFEL if bank == "MIFEL" else SOURCE_BBVA)

    years = None
    if args.pdf_years.strip():
        years = [int(x.strip()) for x in args.pdf_years.split(",") if x.strip()]

    if args.index_pdfs:
        pdf_rows = index_pdfs(args.folder, years)
        print(f"Índice PDF comprobantes: {len(pdf_rows)} archivos parseados por nombre")
        all_records.extend(pdf_rows)
        sources_touched.add(SOURCE_PDF_INDEX)

    if args.index_estados_pdf:
        estado_rows = index_estados_cuenta_pdfs(args.estados_folder, years)
        print(f"Índice PDF estados de cuenta: {len(estado_rows)} archivos en {args.estados_folder}")
        all_records.extend(estado_rows)
        sources_touched.add(SOURCE_ESTADO_CUENTA_PDF_INDEX)

    if not all_records:
        print("Nada que insertar (plantillas vacías o archivos ausentes).")
        if args.dry_run:
            return
        # Still allow wiping nothing
        return

    if args.dry_run:
        print(f"DRY-RUN: {len(all_records)} registros listos, sources={sorted(sources_touched)}")
        for r in all_records[:5]:
            print(" ", r["date"], r["amount"], r["category"], r["description"][:120])
        return

    sb = get_supabase()
    for src in sorted(sources_touched):
        # Preserve manual overrides: fetch existing overridden/obs rows and merge by folio+fecha+bank
        existing = (
            sb.table("financial_records")
            .select("id,date,description,source_file")
            .eq("source_file", src)
            .execute()
        )
        overrides: dict[str, dict] = {}
        for row in existing.data or []:
            try:
                d = json.loads(row.get("description") or "{}")
            except json.JSONDecodeError:
                continue
            if d.get("match_status") == "overridden" or (d.get("observaciones") or "").strip():
                key = "|".join(
                    [
                        str(d.get("bank") or ""),
                        str(d.get("fecha") or row.get("date") or ""),
                        str(d.get("folio") or ""),
                        str(d.get("referencia") or ""),
                        str(d.get("descripcion") or d.get("filename") or ""),
                        str(d.get("cargo") or ""),
                        str(d.get("abono") or d.get("amount") or ""),
                    ]
                )
                overrides[key] = d

        sb.table("financial_records").delete().eq("source_file", src).execute()
        print(f"Limpieza {src}: OK")

        batch_src = [r for r in all_records if r["source_file"] == src]
        if overrides:
            restored = 0
            for r in batch_src:
                try:
                    d = json.loads(r["description"])
                except json.JSONDecodeError:
                    continue
                key = "|".join(
                    [
                        str(d.get("bank") or ""),
                        str(d.get("fecha") or r.get("date") or ""),
                        str(d.get("folio") or ""),
                        str(d.get("referencia") or ""),
                        str(d.get("descripcion") or d.get("filename") or ""),
                        str(d.get("cargo") or ""),
                        str(d.get("abono") or d.get("amount") or ""),
                    ]
                )
                prev = overrides.get(key)
                if not prev:
                    continue
                if prev.get("match_status") == "overridden":
                    d["matched_rubro"] = prev.get("matched_rubro")
                    d["matched_parent"] = prev.get("matched_parent")
                    d["match_status"] = "overridden"
                    d["match_source"] = "manual"
                    d["match_confidence"] = 1.0
                    r["category"] = d["matched_rubro"] or r["category"]
                if (prev.get("observaciones") or "").strip():
                    d["observaciones"] = prev["observaciones"]
                r["description"] = json.dumps(d, ensure_ascii=False)
                restored += 1
            if restored:
                print(f"  Conservados overrides/obs: {restored}")

        inserted = 0
        for batch in chunked(batch_src, 200):
            result = sb.table("financial_records").insert(batch).execute()
            inserted += len(result.data or [])
        print(f"Insertados {src}: {inserted}")


if __name__ == "__main__":
    main()
