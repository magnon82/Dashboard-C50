"""
Generate blank PRESUPUESTO MENSUAL AGOSTO 2026.xlsx machote from July 2026,
using the new RUBRO_CATALOG order and budget formulas.
"""
from __future__ import annotations

import shutil
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

JULY = Path(
    r"I:\.shortcut-targets-by-id\1-6eRRMYs_V7qHEjD8GHjQgwFC63ucMPk"
    r"\PRESUPUESTOS 2026\PRESUPUESTO MENSUAL JULIO 2026.xlsx"
)
DRIVE_2026 = JULY.parent
MI_UNIDAD = Path(r"I:\Mi unidad\Presupuestos")
OUT_NAME = "PRESUPUESTO MENSUAL AGOSTO 2026.xlsx"

N_WEEKS = 5  # Aug 2026: Mon 3→9, 10→16, 17→23, 24→30, 31→Sep 6
SEM_NAMES = [f"SEM {i}" for i in range(1, N_WEEKS + 1)]

# Catalog order (matches app/lib/presupuesto.ts RUBRO_CATALOG)
# kind: parent_cocina | child_cocina | parent_barra | child_barra |
#       parent_servicios | child_servicios | top
ROWS: list[tuple[str, str]] = [
    ("INSUMOS DE COCINA", "parent_cocina"),
    ("Frutas y Verduras", "child_cocina"),
    ("Proteinas", "child_cocina"),
    ("Abarrotes", "child_cocina"),
    ("Lacteos", "child_cocina"),
    ("Panes, tortillas, Postres", "child_cocina"),
    ("Agua", "child_cocina"),
    ("CARBON", "child_cocina"),
    ("INSUMOS DE BARRA", "parent_barra"),
    ("Destilados y vinos", "child_barra"),
    ("Cervezas", "child_barra"),
    ("Abarrotes", "child_barra"),
    ("Café", "child_barra"),
    ("Refrescos, aguas y hielo", "child_barra"),
    ("Frutas y verduras", "child_barra"),
    ("SERVICIOS", "parent_servicios"),
    ("LAVANDERIA", "child_servicios"),
    ("Agua", "child_servicios"),
    ("Gas", "child_servicios"),
    ("Luz", "child_servicios"),
    ("Teléfono", "child_servicios"),
    ("CONTADOR", "child_servicios"),
    ("DISEÑO Y PUBLICIDAD", "child_servicios"),
    ("Alarma", "child_servicios"),
    ("AUDITORIAS", "child_servicios"),
    ("GAS CALENTADORES", "child_servicios"),
    ("MATERIAS PRIMAS", "child_servicios"),
    ("COMIDA PERSONAL", "top"),
    ("RENTA", "top"),
    ("MANTENIMIENTO", "top"),
    ("EQUIPO", "top"),
    ("CRISTALERIA", "top"),
    ("PAPELERIA", "top"),
    ("LIMPIEZA Y BAÑOS", "top"),
    ("GASOLINA Y TAXIS", "top"),
    ("OTROS", "top"),
    ("LICENCIAS Y AFILIACIONES", "top"),
    ("COMISIONES BANCARIAS", "top"),
    ("FINIQUITOS Y RECLUTAMIENTO", "top"),
    ("NÓMINA", "top"),
    ("IMSS", "top"),
    ("IMPUESTOS", "top"),
]

# July template presupuesto amounts for utilities / typical fixed lines
JULY_TEMPLATE_AMOUNTS = {
    "COMIDA PERSONAL": 5000.0,
    "MANTENIMIENTO": 5000.0,
    "PAPELERIA": 1500.0,
    "LIMPIEZA Y BAÑOS": 9600.0,
    "GASOLINA Y TAXIS": 2480.0,
    "OTROS": 1500.0,
    "COMISIONES BANCARIAS": 3000.0,
    # SERVICIOS kids (from July top-level J)
    "Agua|SERVICIOS": 2000.0,
    "Gas|SERVICIOS": 1500.0,
    "Luz|SERVICIOS": 9800.0,
    "Teléfono|SERVICIOS": 1248.0,
    "CONTADOR|SERVICIOS": 6245.0,
    "DISEÑO Y PUBLICIDAD|SERVICIOS": 5000.0,
    "Alarma|SERVICIOS": 610.0,
    "AUDITORIAS|SERVICIOS": 0.0,
    "GAS CALENTADORES|SERVICIOS": 500.0,
    "MATERIAS PRIMAS|SERVICIOS": 4800.0,
}

FILLS = {
    "header": "FF333F4F",
    "parent_cocina": "FF8EAADB",
    "child_cocina": "FFB4C6E7",
    "parent_barra": "FFC8C8C8",
    "child_barra": "FFDADADA",
    "parent_servicios": "FF548135",
    "child_servicios": "FFA9D08E",
    "top": "FF92D050",
    "renta": "FFAEABAB",
    "nomina": "FFFF3300",
    "imss": "FF00B0F0",
    "licencias": "FF757070",
    "gasolina": "FFFF99FF",
    "equipo": "FF99FFCC",
    "papeleria": "FFF4B083",
    "mantenimiento": "FF9CC2E5",
    "limpieza": "FFA8D08D",
    "otros": "FFFFFF99",
    "finiquitos": "FF8EAADB",
}

TOP_FILL_OVERRIDE = {
    "RENTA": "renta",
    "NÓMINA": "nomina",
    "IMSS": "imss",
    "IMPUESTOS": "imss",
    "LICENCIAS Y AFILIACIONES": "licencias",
    "GASOLINA Y TAXIS": "gasolina",
    "EQUIPO": "equipo",
    "CRISTALERIA": "equipo",
    "PAPELERIA": "papeleria",
    "MANTENIMIENTO": "mantenimiento",
    "LIMPIEZA Y BAÑOS": "limpieza",
    "OTROS": "otros",
    "FINIQUITOS Y RECLUTAMIENTO": "finiquitos",
    "COMISIONES BANCARIAS": "finiquitos",
}

FONT_NAME = "Century Gothic"
THIN = Border(
    left=Side(style="thin", color="FFB0B0B0"),
    right=Side(style="thin", color="FFB0B0B0"),
    top=Side(style="thin", color="FFB0B0B0"),
    bottom=Side(style="thin", color="FFB0B0B0"),
)


def fill_of(rgb: str) -> PatternFill:
    return PatternFill(start_color=rgb, end_color=rgb, fill_type="solid")


def style_label(cell, *, bold: bool, fill_key: str, white_font: bool = False):
    cell.font = Font(
        name=FONT_NAME,
        size=11,
        bold=bold,
        color="FFFFFFFF" if white_font else "FF000000",
    )
    cell.fill = fill_of(FILLS[fill_key])
    cell.alignment = Alignment(vertical="center")


def style_num(cell, *, fill_key: str | None = None):
    cell.font = Font(name=FONT_NAME, size=11)
    if fill_key:
        cell.fill = fill_of(FILLS[fill_key])
    cell.number_format = "#,##0.00"
    cell.alignment = Alignment(horizontal="right", vertical="center")


def clear_cell(cell):
    cell.value = None


def sem_sum(col: str, row: int) -> str:
    parts = [f"'{name}'!{col}{row}" for name in SEM_NAMES]
    return "=" + "+".join(parts)


def row_map() -> dict[int, tuple[str, str]]:
    """1-based Excel row -> (name, kind). Data starts at row 2."""
    return {i + 2: ROWS[i] for i in range(len(ROWS))}


def fill_key_for(name: str, kind: str) -> str:
    if kind.startswith("parent_") or kind.startswith("child_"):
        return kind
    return TOP_FILL_OVERRIDE.get(name, "top")


def presupuesto_value(name: str, kind: str, row: int, rows: dict) -> object | None:
    """Return J-column value/formula, or None if no presupuesto on this row."""
    # Parents
    if kind == "parent_cocina":
        return "=$B$49*16%"
    if kind == "parent_barra":
        return "=$B$49*14%"
    if kind == "parent_servicios":
        # kids are rows after parent until next non-child_servicios
        start = row + 1
        end = row
        for r, (_n, k) in rows.items():
            if r > row and k == "child_servicios":
                end = r
        return f"=SUM(J{start}:J{end})"

    if kind == "child_cocina" and name == "CARBON":
        return 1500 * N_WEEKS
    if kind in ("child_cocina", "child_barra"):
        return None  # only parent % (CARBON handled above)

    if kind == "child_servicios":
        if name == "LAVANDERIA":
            return 2400 * N_WEEKS
        key = f"{name}|SERVICIOS"
        return JULY_TEMPLATE_AMOUNTS.get(key, 0.0)

    # top-level formulas / defaults
    if name == "RENTA":
        return 44330
    if name == "CRISTALERIA":
        return 500
    if name == "EQUIPO":
        return 5000
    if name == "LICENCIAS Y AFILIACIONES":
        return 3500
    if name == "FINIQUITOS Y RECLUTAMIENTO":
        return 0
    if name == "NÓMINA":
        return "=$B$49*25%"
    if name == "IMSS":
        return 16765.12
    if name == "IMPUESTOS":
        return 6000
    return JULY_TEMPLATE_AMOUNTS.get(name, 0.0)


def has_ijk(name: str, kind: str) -> bool:
    """Whether row gets % / presupuesto / real columns on TOTAL."""
    if kind in ("parent_cocina", "parent_barra", "parent_servicios", "top"):
        return True
    if kind == "child_cocina" and name == "CARBON":
        return True  # J only — handled specially
    if kind == "child_servicios":
        return True  # J (+ optional I/K); we set J and skip I/K to avoid double-count
    return False


def parent_and_top_rows(rows: dict) -> list[int]:
    return [
        r
        for r, (_n, k) in rows.items()
        if k in ("parent_cocina", "parent_barra", "parent_servicios", "top")
    ]


def clear_rubro_block(ws, max_clear_row: int = 46):
    for r in range(2, max_clear_row + 1):
        for c in range(1, 12):  # A-K
            cell = ws.cell(r, c)
            # keep L (SEM labels) on TOTAL
            clear_cell(cell)
            if c in (1, 4, 7):
                cell.fill = PatternFill(fill_type=None)
                cell.font = Font(name=FONT_NAME, size=11)


def unmerge_in_rows(ws, min_row: int, max_row: int):
    to_remove = []
    for mr in list(ws.merged_cells.ranges):
        if mr.min_row >= min_row and mr.max_row <= max_row:
            to_remove.append(str(mr))
        elif mr.min_row <= max_row and mr.max_row >= min_row:
            to_remove.append(str(mr))
    for ref in to_remove:
        try:
            ws.unmerge_cells(ref)
        except Exception:
            pass


def write_sem_sheet(ws, rows: dict):
    unmerge_in_rows(ws, 1, 50)
    # Header row 1 labels kept / refreshed
    for col, title in ((1, "GASTOS EN EFECTIVO"), (4, "GASTOS BANCO"), (7, "GASTOS BANCO")):
        cell = ws.cell(1, col, title)
        style_label(cell, bold=True, fill_key="header", white_font=True)

    clear_rubro_block(ws, 50)

    sum_rows = parent_and_top_rows(rows)
    for col in ("B", "E", "H"):
        formula = "=" + "+".join(f"{col}{r}" for r in sum_rows)
        cell = ws[f"{col}1"]
        cell.value = formula
        style_num(cell, fill_key="header")
        cell.font = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFFFF")

    for r, (name, kind) in rows.items():
        fk = fill_key_for(name, kind)
        bold = kind.startswith("parent_") or kind == "top"
        for col in (1, 4, 7):
            cell = ws.cell(r, col, name)
            style_label(cell, bold=bold, fill_key=fk, white_font=(fk == "header"))

        # Parent SUM formulas
        if kind == "parent_cocina":
            for col in ("B", "E", "H"):
                cell = ws[f"{col}{r}"]
                cell.value = f"=SUM({col}{r+1}:{col}{r+7})"
                style_num(cell, fill_key=fk)
        elif kind == "parent_barra":
            for col in ("B", "E", "H"):
                cell = ws[f"{col}{r}"]
                cell.value = f"=SUM({col}{r+1}:{col}{r+6})"
                style_num(cell, fill_key=fk)
        elif kind == "parent_servicios":
            for col in ("B", "E", "H"):
                cell = ws[f"{col}{r}"]
                cell.value = f"=SUM({col}{r+1}:{col}{r+11})"
                style_num(cell, fill_key=fk)
        else:
            for col in ("B", "E", "H"):
                cell = ws[f"{col}{r}"]
                cell.value = 0
                style_num(cell, fill_key=fk)

        # Clear note columns C, F, I
        for col in (3, 6, 9):
            clear_cell(ws.cell(r, col))

    last = max(rows)
    # Channel footers (like July)
    ws.cell(last + 1, 1, "EFECTIVO")
    style_label(ws.cell(last + 1, 1), bold=True, fill_key="header", white_font=True)
    ws.cell(last + 2, 4, "MIFEL")
    style_label(ws.cell(last + 2, 4), bold=True, fill_key="header", white_font=True)
    ws.cell(last + 2, 7, "BBVA")
    style_label(ws.cell(last + 2, 7), bold=True, fill_key="header", white_font=True)
    try:
        ws.merge_cells(start_row=last + 2, start_column=4, end_row=last + 3, end_column=4)
        ws.merge_cells(start_row=last + 2, start_column=7, end_row=last + 3, end_column=7)
    except Exception:
        pass


def write_total_sheet(ws, rows: dict):
    unmerge_in_rows(ws, 2, 46)
    # Ensure header titles
    ws["A1"] = "GASTOS EN EFECTIVO"
    ws["D1"] = "GASTOS BANCO MIFEL"
    ws["G1"] = "GASTOS BANCO BBVA"
    ws["J1"] = "PRESUPUESTO"
    style_label(ws["A1"], bold=True, fill_key="header", white_font=True)
    style_label(ws["D1"], bold=True, fill_key="header", white_font=True)
    style_label(ws["G1"], bold=True, fill_key="header", white_font=True)
    style_label(ws["J1"], bold=True, fill_key="header", white_font=True)

    clear_rubro_block(ws, 46)

    sum_rows = parent_and_top_rows(rows)
    for col in ("B", "E", "H"):
        formula = "=" + "+".join(f"{col}{r}" for r in sum_rows)
        cell = ws[f"{col}1"]
        cell.value = formula
        style_num(cell, fill_key="header")
        cell.font = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFFFF")

    for r, (name, kind) in rows.items():
        fk = fill_key_for(name, kind)
        bold = kind.startswith("parent_") or kind == "top"
        for col in (1, 4, 7):
            cell = ws.cell(r, col, name)
            style_label(cell, bold=bold, fill_key=fk)

        if kind == "parent_cocina":
            for col in ("B", "E", "H"):
                cell = ws[f"{col}{r}"]
                cell.value = f"=SUM({col}{r+1}:{col}{r+7})"
                style_num(cell, fill_key=fk)
        elif kind == "parent_barra":
            for col in ("B", "E", "H"):
                cell = ws[f"{col}{r}"]
                cell.value = f"=SUM({col}{r+1}:{col}{r+6})"
                style_num(cell, fill_key=fk)
        elif kind == "parent_servicios":
            for col in ("B", "E", "H"):
                cell = ws[f"{col}{r}"]
                cell.value = f"=SUM({col}{r+1}:{col}{r+11})"
                style_num(cell, fill_key=fk)
        else:
            for col in ("B", "E", "H"):
                cell = ws[f"{col}{r}"]
                cell.value = sem_sum(col, r)
                style_num(cell, fill_key=fk)

        # I / J / K
        j_val = presupuesto_value(name, kind, r, rows)
        show_full_ijk = kind in (
            "parent_cocina",
            "parent_barra",
            "parent_servicios",
            "top",
        )
        show_j_only = (kind == "child_servicios") or (
            kind == "child_cocina" and name == "CARBON"
        )

        if show_full_ijk:
            ws[f"K{r}"] = f"=B{r}+E{r}+H{r}"
            style_num(ws[f"K{r}"])
            ws[f"I{r}"] = f"=IF($B$49=0,0,K{r}/$B$49)"
            ws[f"I{r}"].number_format = "0.00%"
            ws[f"I{r}"].font = Font(name=FONT_NAME, size=11)
            if j_val is not None:
                ws[f"J{r}"] = j_val
                style_num(ws[f"J{r}"])
        elif show_j_only and j_val is not None:
            ws[f"J{r}"] = j_val
            style_num(ws[f"J{r}"])
            # no I/K on nested kids (avoids double-count in totals)

    # Footer meta (same rows as July)
    ws["A47"] = "efe"
    ws["B47"] = 0
    style_num(ws["B47"])
    ws["A48"] = "ba"
    ws["B48"] = "=M2+M21"
    style_num(ws["B48"])
    ws["A49"] = "venta"
    ws["B49"] = 0
    style_num(ws["B49"])

    # Total real spend / balance formulas — parents + tops only
    k_parts = "+".join(f"K{r}" for r in sum_rows)
    ws["K47"] = f"={k_parts}"
    style_num(ws["K47"])
    ws["B51"] = "=B47+B48+B49"
    style_num(ws["B51"])
    ws["K50"] = "=B51-K47"
    style_num(ws["K50"])

    # Clear leftover old rubro labels below last data row up to 46
    last = max(rows)
    for r in range(last + 1, 47):
        for c in range(1, 12):
            if r in (47,) or (r >= 47):
                break
            clear_cell(ws.cell(r, c))


def apply_budget_conditional_formatting(ws, first_row: int = 2, last_row: int = 43):
    """
    TOTAL: alert Real (K) and % (I) vs Presupuesto (J).
    Red when Real > Presupuesto; green when Real ≤ Presupuesto.
    SEM sheets have no J/K presupuesto block — CF not applicable there.
    """
    from openpyxl.formatting.formatting import ConditionalFormattingList

    ws.conditional_formatting = ConditionalFormattingList()

    red = PatternFill(start_color="FFFF6B6B", end_color="FFFF6B6B", fill_type="solid")
    green = PatternFill(start_color="FF92D050", end_color="FF92D050", fill_type="solid")

    over = f'AND(ISNUMBER($J{first_row}),ISNUMBER($K{first_row}),$K{first_row}>$J{first_row})'
    within = f'AND(ISNUMBER($J{first_row}),ISNUMBER($K{first_row}),$K{first_row}<=$J{first_row})'

    for col in ("K", "I"):
        rng = f"{col}{first_row}:{col}{last_row}"
        ws.conditional_formatting.add(rng, FormulaRule(formula=[over], fill=red))
        ws.conditional_formatting.add(rng, FormulaRule(formula=[within], fill=green))


def zero_bank_panel(ws):
    """Blank machote: zero inputs, keep roll-forward formulas."""
    # Mifel inicial / weekly ventas / anticipos
    for coord in [
        "M2",
        "M5",
        "M6",
        "M7",
        "M8",
        "M9",
        "M11",
        "M12",
        "M13",
        "M14",
        "M15",
        "M21",
        "M24",
        "M25",
        "M26",
        "M27",
        "M28",
        "M30",
        "M31",
        "M32",
        "M33",
        "M34",
        "O2",
        "O21",
    ]:
        cell = ws[coord]
        # keep formulas (M16, M35, N2, etc.)
        if isinstance(cell.value, str) and str(cell.value).startswith("="):
            continue
        cell.value = 0

    for r in range(5, 10):
        for col in ("N", "O", "P"):
            cell = ws[f"{col}{r}"]
            if isinstance(cell.value, str) and str(cell.value).startswith("="):
                continue
            cell.value = 0
    for r in range(24, 29):
        for col in ("N", "O", "P"):
            cell = ws[f"{col}{r}"]
            if isinstance(cell.value, str) and str(cell.value).startswith("="):
                continue
            cell.value = 0
    for r in range(11, 16):
        cell = ws[f"N{r}"]
        if isinstance(cell.value, str) and str(cell.value).startswith("="):
            continue
        cell.value = 0


def validate(path: Path) -> list[str]:
    notes = []
    wb = load_workbook(path, data_only=False)
    expected = ["SEM 1", "SEM 2", "SEM 3", "SEM 4", "SEM 5", "TOTAL"]
    if wb.sheetnames != expected:
        notes.append(f"Sheets={wb.sheetnames} (expected {expected})")
    ws = wb["TOTAL"]
    rows = row_map()
    for r, (name, _k) in rows.items():
        if ws.cell(r, 1).value != name:
            notes.append(f"A{r}={ws.cell(r,1).value!r} expected {name!r}")
    # spot-check formulas
    for coord in ["B1", "B2", "B10", "B17", "J2", "J41", "K47", "B49"]:
        v = ws[coord].value
        if v is None:
            notes.append(f"{coord} empty")
        if isinstance(v, str) and "#REF!" in v:
            notes.append(f"{coord} has #REF!")
    # no #REF! anywhere in TOTAL A-K
    for row in ws.iter_rows(min_row=1, max_row=55, max_col=12):
        for c in row:
            if isinstance(c.value, str) and "#REF!" in c.value:
                notes.append(f"{c.coordinate} #REF!")
    wb.close()
    return notes


def main():
    if not JULY.exists():
        raise SystemExit(f"July file not found: {JULY}")

    MI_UNIDAD.mkdir(parents=True, exist_ok=True)
    out_drive = DRIVE_2026 / OUT_NAME
    out_mi = MI_UNIDAD / OUT_NAME

    # Clone July (preserves styles, bank panels, SEM sheets)
    shutil.copy2(JULY, out_drive)

    wb = load_workbook(out_drive)
    rows = row_map()

    # Ensure SEM 1..5 exist (July already has them)
    for name in SEM_NAMES:
        if name not in wb.sheetnames:
            raise SystemExit(f"Missing sheet {name}")

    for name in SEM_NAMES:
        write_sem_sheet(wb[name], rows)

    write_total_sheet(wb["TOTAL"], rows)
    zero_bank_panel(wb["TOTAL"])
    apply_budget_conditional_formatting(wb["TOTAL"], first_row=2, last_row=1 + len(ROWS))

    # Re-order sheets: SEM 1..N then TOTAL (like July)
    order = SEM_NAMES + ["TOTAL"]
    for i, name in enumerate(order):
        wb.move_sheet(name, offset=i - wb.sheetnames.index(name))

    wb.save(out_drive)
    wb.close()

    # Also place next to Mi unidad\Presupuestos shortcuts
    shutil.copy2(out_drive, out_mi)

    notes = validate(out_drive)
    print("OK", out_drive)
    print("COPY", out_mi)
    print("N_WEEKS", N_WEEKS)
    print("ROWS", len(ROWS), "excel rows 2-" + str(1 + len(ROWS)))
    if notes:
        print("VALIDATION NOTES:")
        for n in notes:
            print(" -", n)
    else:
        print("VALIDATION: ok")


if __name__ == "__main__":
    main()
