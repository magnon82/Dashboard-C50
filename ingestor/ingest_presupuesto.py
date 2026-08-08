"""
Ingestor: PRESUPUESTO MENSUAL (hoja TOTAL + SEM n) → financial_records.

source_file:
  - presupuesto_mensual  — gastos por canal (Efectivo/Mifel/BBVA) por rubro
  - presupuesto_saldos   — saldo ACTUAL Mifel / BBVA del mes
  - presupuesto_rubro    — presupuesto vs real por rubro (+ canales, padre)
  - presupuesto_semana   — resumen semanal de movimientos (TOTAL!U:Z)
  - presupuesto_sem_detalle — gasto por SEM × rubro × canal (+ nota de concepto)
  - presupuesto_ingreso  — ingresos bancarios semanales Mifel/BBVA (TOTAL, llenado manual)
"""

from __future__ import annotations

import argparse
import json
import os
import re
import unicodedata
from datetime import date, timedelta
from pathlib import Path

from dotenv import load_dotenv
from openpyxl import load_workbook
from supabase import create_client

load_dotenv()
load_dotenv(Path(__file__).resolve().parent.parent / ".env.local")

SOURCE_FILE = "presupuesto_mensual"
SOURCE_SALDOS = "presupuesto_saldos"
SOURCE_RUBRO = "presupuesto_rubro"
SOURCE_SEMANA = "presupuesto_semana"
SOURCE_DETALLE = "presupuesto_sem_detalle"
SOURCE_INGRESO = "presupuesto_ingreso"
# Admin overrides live in the app (source_file=presupuesto_ajuste); not wiped here.

DEFAULT_YEAR_FOLDER = Path(
    r"I:\.shortcut-targets-by-id\1-6eRRMYs_V7qHEjD8GHjQgwFC63ucMPk\PRESUPUESTOS 2026"
)

YEAR_FOLDERS = [
    Path(r"I:\.shortcut-targets-by-id\1dDDnlR8VfbCeaI1Hn0cPg7HBHqfyUJFA\PRESUPUESTOS 2022"),
    Path(r"I:\.shortcut-targets-by-id\1c6J44HPdUaoGKQI8RcRdBtxg-c0HZMAT\PRESUPUESTOS 2023"),
    Path(
        r"I:\.shortcut-targets-by-id\1-2gKQvuVI_3O2N5-uZ2NG51FbQftMqSG\PRESUPUESTOS MENSUALES  2024"
    ),
    Path(
        r"I:\.shortcut-targets-by-id\10X4rPJGf3mGqVWGybos3O11nHD_4e7Cx\PRESUPUESTOS MENSUALES 2025"
    ),
    DEFAULT_YEAR_FOLDER,
]

MONTHS = {
    "ENERO": 1,
    "FEBRERO": 2,
    "FEBRRERO": 2,
    "MARZO": 3,
    "ABRIL": 4,
    "MAYO": 5,
    "JUNIO": 6,
    "JULIO": 7,
    "AGOSTO": 8,
    "SEPTIEMBRE": 9,
    "OCTUBRE": 10,
    "NOVIEMBRE": 11,
    "DICIEMBRE": 12,
}

PARENT_COCINA = "Insumos de cocina"
PARENT_BARRA = "Insumos de barra"
PARENT_SERVICIOS = "Servicios"

# Keys (norm / upper) → display names for stored parent fields
PARENT_DISPLAY = {
    "INSUMOS DE COCINA": PARENT_COCINA,
    "INSUMOS DE BARRA": PARENT_BARRA,
    "SERVICIOS": PARENT_SERVICIOS,
}
PARENT_CATEGORIES = set(PARENT_DISPLAY.keys())

COCINA_CHILDREN = {
    "FRUTAS Y VERDURAS",
    "PROTEINAS",
    "ABARROTES",
    "LACTEOS",
    "PANES, TORTILLAS, POSTRES",
    "AGUA",
    "CARBON",
}

BARRA_CHILDREN = {
    "DESTILADOS Y VINOS",
    "CERVEZAS",
    "ABARROTES",
    "CAFE",
    "CAFÉ",
    "REFRESCOS, AGUAS Y HIELO",
    "FRUTAS Y VERDURAS",
}

SERVICIOS_CHILDREN = {
    "LAVANDERIA",
    "AGUA",
    "GAS",
    "LUZ",
    "TELEFONO",
    "TELÉFONO",
    "CONTADOR",
    "DISEÑO Y PUBLICIDAD",
    "DISENO Y PUBLICIDAD",
    "ALARMA",
    "AUDITORIAS",
    "GAS CALENTADORES",
    "MATERIAS PRIMAS",
}

SKIP_NAMES = {
    "GASTOS EN EFECTIVO",
    "GASTOS BANCO MIFEL",
    "GASTOS BANCO BBVA",
    "GASTOS BANCO",
    "EFE",
    "BA",
    "VENTA",
    "INICIAL EFE",
    "INICIAL BANCOS",
    "FINAL BANCOS",
    "FINAL EFE",
}

STREAMS = (
    (0, 1, "Efectivo"),
    (3, 4, "Mifel"),
    (6, 7, "BBVA"),
)

ALL_SOURCES = (
    SOURCE_FILE,
    SOURCE_SALDOS,
    SOURCE_RUBRO,
    SOURCE_SEMANA,
    SOURCE_DETALLE,
    SOURCE_INGRESO,
)


def first_monday_on_or_after(year: int, month: int, day: int) -> date:
    d = date(year, month, day)
    while d.weekday() != 0:
        d += timedelta(days=1)
    return d


def monday_of_month_sem(year: int, month: int, week: int) -> date:
    """Lunes de SEM n del mes (alineado a presupuesto / flujo efectivo)."""
    return first_monday_on_or_after(year, month, 1) + timedelta(days=(week - 1) * 7)


def parse_month_year(filename: str) -> tuple[int, int] | None:
    upper = filename.upper()
    year_m = re.search(r"(20\d{2})", upper)
    if not year_m:
        return None
    year = int(year_m.group(1))
    for name, num in MONTHS.items():
        if name in upper:
            return year, num
    return None


def as_amount(value) -> float | None:
    if value is None or isinstance(value, str):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def money_or_0(value) -> float:
    v = as_amount(value)
    return 0.0 if v is None else float(v)


def norm_cat(name: str) -> str:
    return re.sub(r"\s+", " ", name.strip())


def norm_rubro_key(name: str) -> str:
    """Match frontend: strip accents, upper, collapse non-alnum."""
    nfkd = unicodedata.normalize("NFD", name)
    plain = "".join(c for c in nfkd if unicodedata.category(c) != "Mn")
    return re.sub(r"[^A-Z0-9]+", " ", plain.upper()).strip()


SERVICIOS_CHILDREN_KEYS = {norm_rubro_key(n) for n in SERVICIOS_CHILDREN}


def rubro_merge_key(rubro: str, parent: str | None) -> str:
    p = norm_rubro_key(parent) if parent else ""
    return f"{p}::{norm_rubro_key(rubro)}"


def parent_section_key(name: str | None) -> str:
    """Normalized key for parent section comparisons."""
    if not name:
        return ""
    return norm_rubro_key(name)


def parent_display(name_or_key: str) -> str:
    key = norm_rubro_key(name_or_key)
    return PARENT_DISPLAY.get(key, name_or_key)


# Legacy Excel rubros → single catalog line (matches app/lib/presupuesto.ts)
RUBRO_CRISTALERIA_Y_EQUIPO = "Cristalería y Equipo"
_LEGACY_CRISTALERIA_EQUIPO = {"EQUIPO", "CRISTALERIA", "CRISTALERIA Y EQUIPO"}

# Display names keyed by norm_rubro_key (accents stripped, upper)
RUBRO_DISPLAY_BY_KEY = {
    "INSUMOS DE COCINA": PARENT_COCINA,
    "INSUMOS DE BARRA": PARENT_BARRA,
    "SERVICIOS": PARENT_SERVICIOS,
    "FRUTAS Y VERDURAS": "Frutas y verduras",
    "PROTEINAS": "Proteínas",
    "ABARROTES": "Abarrotes",
    "LACTEOS": "Lácteos",
    "PANES TORTILLAS POSTRES": "Panes, tortillas, postres",
    "AGUA": "Agua",
    "CARBON": "Carbón",
    "DESTILADOS Y VINOS": "Destilados y vinos",
    "CERVEZAS": "Cervezas",
    "CAFE": "Café",
    "REFRESCOS AGUAS Y HIELO": "Refrescos, aguas y hielo",
    "LAVANDERIA": "Lavandería",
    "GAS": "Gas",
    "LUZ": "Luz",
    "TELEFONO": "Teléfono",
    "CONTADOR": "Contador",
    "DISENO Y PUBLICIDAD": "Diseño y publicidad",
    "ALARMA": "Alarma",
    "AUDITORIAS": "Auditorías",
    "GAS CALENTADORES": "Gas calentadores",
    "MATERIAS PRIMAS": "Materias primas",
    "COMIDA PERSONAL": "Comida personal",
    "RENTA": "Renta",
    "MANTENIMIENTO": "Mantenimiento",
    "CRISTALERIA Y EQUIPO": RUBRO_CRISTALERIA_Y_EQUIPO,
    "PAPELERIA": "Papelería",
    "LIMPIEZA Y BANOS": "Limpieza y baños",
    "GASOLINA Y TAXIS": "Gasolina y taxis",
    "OTROS": "Otros",
    "LICENCIAS Y AFILIACIONES": "Licencias y afiliaciones",
    "COMISIONES BANCARIAS": "Comisiones bancarias",
    "FINIQUITOS Y RECLUTAMIENTO": "Finiquitos y reclutamiento",
    "NOMINA": "Nómina",
    "IMSS": "IMSS",
    "IMPUESTOS": "Impuestos",
}


def canonicalize_rubro_name(rubro: str) -> str:
    key = norm_rubro_key(rubro)
    if key in _LEGACY_CRISTALERIA_EQUIPO:
        return RUBRO_CRISTALERIA_Y_EQUIPO
    return RUBRO_DISPLAY_BY_KEY.get(key, rubro)


def months_in_recent_window(days: int, today: date | None = None) -> set[tuple[int, int]]:
    """Year/month pairs overlapping the last `days` calendar days."""
    today = today or date.today()
    start = today - timedelta(days=max(days, 1))
    out: set[tuple[int, int]] = set()
    cur = date(start.year, start.month, 1)
    end = date(today.year, today.month, 1)
    while cur <= end:
        out.add((cur.year, cur.month))
        if cur.month == 12:
            cur = date(cur.year + 1, 1, 1)
        else:
            cur = date(cur.year, cur.month + 1, 1)
    return out


def extract_saldos_from_total(rows: list, year: int, month: int, month_label: str) -> list[dict]:
    if len(rows) < 2:
        return []

    month_date = f"{year:04d}-{month:02d}-01"
    records: list[dict] = []

    r2 = list(rows[1]) + [None] * 19
    mifel = as_amount(r2[14])  # O
    if mifel is not None and mifel != 0:
        records.append(
            {
                "date": month_date,
                "type": "income",
                "category": "Saldo Mifel",
                "amount": abs(mifel),
                "description": f"{month_label} · MIFEL ACTUAL",
                "source_file": SOURCE_SALDOS,
            }
        )

    bbva = None
    for i, row in enumerate(rows):
        cells = list(row) + [None] * 19
        label = str(cells[14] or "").strip().upper()
        if label == "BBVA":
            for j in range(i + 1, min(i + 6, len(rows))):
                c2 = list(rows[j]) + [None] * 19
                val = as_amount(c2[14])
                if val is not None:
                    bbva = val
                    break
            break

    if bbva is not None and bbva != 0:
        records.append(
            {
                "date": month_date,
                "type": "income",
                "category": "Saldo BBVA",
                "amount": abs(bbva),
                "description": f"{month_label} · BBVA ACTUAL",
                "source_file": SOURCE_SALDOS,
            }
        )

    return records


def detect_parent(rubro: str, row_idx: int, last_parent: str | None) -> tuple[str | None, str | None]:
    """Returns (parent_for_this_row, new_last_parent). Parents are display names."""
    key = norm_rubro_key(rubro)
    upper = rubro.upper()
    last_key = parent_section_key(last_parent)

    if key in PARENT_DISPLAY:
        return None, PARENT_DISPLAY[key]
    if last_key == "INSUMOS DE COCINA":
        if upper in COCINA_CHILDREN or upper.startswith("PANES") or key == "CARBON":
            return last_parent, last_parent
        if key in PARENT_DISPLAY:
            return None, PARENT_DISPLAY[key]
        # left cocina block when we hit a top-level non-child
        if upper not in BARRA_CHILDREN:
            return None, None
    if last_key == "INSUMOS DE BARRA":
        if upper in BARRA_CHILDREN or upper in {"CAFE", "CAFÉ"}:
            return last_parent, last_parent
        return None, None
    if last_key == "SERVICIOS":
        if upper in SERVICIOS_CHILDREN or key in SERVICIOS_CHILDREN_KEYS:
            return last_parent, last_parent
        return None, None
    # Flat Excel names remapped by frontend; optional ingest hints:
    if key == "CARBON":
        return PARENT_COCINA, last_parent
    if key in SERVICIOS_CHILDREN_KEYS and key != "AGUA":
        # Agua is ambiguous (cocina vs servicios); leave flat for frontend remap
        return PARENT_SERVICIOS, last_parent
    return None, last_parent


def extract_rubros(
    rows: list, year: int, month: int, month_label: str
) -> tuple[list[dict], list[dict], dict]:
    """Channel expenses + rubro resumen. Also returns meta totals."""
    month_date = f"{year:04d}-{month:02d}-01"
    channel_rows: list[dict] = []
    rubro_rows: list[dict] = []
    meta = {"venta": 0.0, "efe": 0.0, "ba": 0.0}

    last_parent: str | None = None
    seen_barra_parent = False

    for i, row in enumerate(rows[:55]):
        if i == 0:
            continue
        cells = list(row) + [None] * 12

        # Footer meta
        a = str(cells[0] or "").strip().lower()
        if a in ("venta",):
            meta["venta"] = money_or_0(cells[1])
        elif a in ("efe", "inicial efe"):
            meta["efe"] = money_or_0(cells[1])
        elif a in ("ba", "inicial bancos"):
            meta["ba"] = money_or_0(cells[1])
        elif "final bancos" in str(cells[9] or "").strip().lower():
            meta["ba"] = money_or_0(cells[10])

        # Resolve rubro name: prefer Mifel label, then Efectivo, then BBVA
        name_m = cells[3]
        name_e = cells[0]
        name_b = cells[6]
        raw = name_m or name_e or name_b
        if not raw:
            continue
        rubro = norm_cat(str(raw))
        upper = rubro.upper()
        if upper in SKIP_NAMES:
            continue

        # Track parent sections by row order on Efectivo/Mifel columns
        e_name = norm_cat(str(name_e)) if name_e else ""
        e_upper = e_name.upper()
        e_key = norm_rubro_key(e_name) if e_name else ""
        last_key = parent_section_key(last_parent)
        if e_key in PARENT_DISPLAY:
            last_parent = PARENT_DISPLAY[e_key]
            if e_key == "INSUMOS DE BARRA":
                seen_barra_parent = True
            parent = None
            is_parent = True
            rubro = PARENT_DISPLAY[e_key]
        elif last_key == "INSUMOS DE COCINA" and (
            e_upper in COCINA_CHILDREN
            or e_upper.startswith("PANES")
            or e_key == "CARBON"
        ):
            parent = PARENT_COCINA
            is_parent = False
            rubro = e_name or rubro
        elif last_key == "INSUMOS DE BARRA" and (
            e_upper in BARRA_CHILDREN or e_upper in {"CAFE", "CAFÉ"} or upper in BARRA_CHILDREN
        ):
            parent = PARENT_BARRA
            is_parent = False
            # Prefer efectivo name when present for barra children
            rubro = e_name or rubro
        elif last_key == "SERVICIOS" and (
            e_upper in SERVICIOS_CHILDREN or e_key in SERVICIOS_CHILDREN_KEYS
        ):
            parent = PARENT_SERVICIOS
            is_parent = False
            rubro = e_name or rubro
        else:
            if e_upper and e_upper not in COCINA_CHILDREN and e_upper not in BARRA_CHILDREN:
                if last_key == "INSUMOS DE COCINA" and e_key != "INSUMOS DE BARRA":
                    # still in cocina until barra parent or other major section
                    if e_upper.startswith("INSUMOS") or e_key.startswith("INSUMOS"):
                        last_parent = PARENT_DISPLAY.get(e_key)
                    elif money_or_0(cells[9]) or money_or_0(cells[10]) or name_e:
                        # top-level rubro (Comida personal, etc.)
                        last_parent = None
            parent = None
            is_parent = e_key in PARENT_DISPLAY or upper in PARENT_CATEGORIES
            if is_parent:
                last_parent = PARENT_DISPLAY.get(e_key) or PARENT_DISPLAY.get(
                    norm_rubro_key(rubro), rubro
                )
                rubro = last_parent or rubro
            # Flat remaps (frontend catalog is source of truth for display)
            elif e_key == "CARBON" or norm_rubro_key(rubro) == "CARBON":
                parent = PARENT_COCINA
                is_parent = False
            elif norm_rubro_key(rubro) in {
                "LAVANDERIA",
                "CONTADOR",
                "DISENO Y PUBLICIDAD",
                "AUDITORIAS",
                "GAS CALENTADORES",
                "MATERIAS PRIMAS",
            }:
                parent = PARENT_SERVICIOS
                is_parent = False

        efectivo = money_or_0(cells[1]) if name_e else money_or_0(cells[1]) if cells[1] else 0.0
        mifel = money_or_0(cells[4])
        bbva = money_or_0(cells[7])
        # If only Mifel/BBVA columns have the name for this row
        if name_e and not name_m and not name_b:
            mifel = 0.0
            bbva = 0.0
            efectivo = money_or_0(cells[1])
        else:
            efectivo = money_or_0(cells[1]) if name_e else 0.0
            mifel = money_or_0(cells[4]) if name_m else 0.0
            bbva = money_or_0(cells[7]) if name_b else 0.0
            # When names repeat across streams on same row, take all three
            if name_e and name_m and name_b:
                efectivo = money_or_0(cells[1])
                mifel = money_or_0(cells[4])
                bbva = money_or_0(cells[7])

        presupuesto = money_or_0(cells[9])  # J
        real_k = money_or_0(cells[10])  # K
        pct = as_amount(cells[8])  # I
        real = real_k if real_k else (efectivo + mifel + bbva)

        # Skip empty noise rows (keep children under a parent even if 0)
        if (
            not is_parent
            and parent is None
            and efectivo == 0
            and mifel == 0
            and bbva == 0
            and presupuesto == 0
            and real == 0
        ):
            continue
        if not is_parent and parent and efectivo == 0 and mifel == 0 and bbva == 0 and real == 0:
            # still keep known zero children for expand UI
            pass
        payload = {
            "rubro": rubro,
            "efectivo": efectivo,
            "mifel": mifel,
            "bbva": bbva,
            "presupuesto": presupuesto,
            "real": real,
            "pct": pct,
            "parent": parent,
            "isParent": is_parent,
            "sort": i,
        }
        rubro_rows.append(payload)

        if not is_parent:
            for canal, amount in (
                ("Efectivo", efectivo),
                ("Mifel", mifel),
                ("BBVA", bbva),
            ):
                if amount == 0:
                    continue
                channel_rows.append(
                    {
                        "date": month_date,
                        "type": "expense",
                        "category": f"{canal}: {rubro}",
                        "amount": amount,
                        "description": f"{month_label} · {canal} · {rubro}",
                        "source_file": SOURCE_FILE,
                    }
                )

    # Merge Excel split rows (same rubro across Efectivo/Mifel/BBVA) → 1 row
    # Also fold legacy EQUIPO + CRISTALERIA into Cristalería y Equipo
    merged: dict[str, dict] = {}
    for p in rubro_rows:
        p = dict(p)
        p["rubro"] = canonicalize_rubro_name(str(p["rubro"]))
        if p.get("parent"):
            p["parent"] = parent_display(str(p["parent"]))
        key = rubro_merge_key(p["rubro"], p["parent"])
        if key not in merged:
            merged[key] = dict(p)
            continue
        cur = merged[key]
        cur["efectivo"] = float(cur["efectivo"]) + float(p["efectivo"])
        cur["mifel"] = float(cur["mifel"]) + float(p["mifel"])
        cur["bbva"] = float(cur["bbva"]) + float(p["bbva"])
        cur["presupuesto"] = float(cur["presupuesto"]) + float(p["presupuesto"])
        cur["real"] = float(cur["real"]) + float(p["real"])
        if p.get("pct") is not None and cur.get("pct") is None:
            cur["pct"] = p["pct"]
        cur["sort"] = min(int(cur["sort"]), int(p["sort"]))
        if p.get("isParent"):
            cur["isParent"] = True
            cur["parent"] = None

    out_rubros: list[dict] = []
    for p in sorted(merged.values(), key=lambda x: int(x["sort"])):
        real = float(p["real"])
        if not real:
            real = float(p["efectivo"]) + float(p["mifel"]) + float(p["bbva"])
            p["real"] = real
        out_rubros.append(
            {
                "date": month_date,
                "type": "expense",
                "category": p["rubro"],
                "amount": real,
                "description": json.dumps(p, ensure_ascii=False),
                "source_file": SOURCE_RUBRO,
            }
        )

    # Month meta row
    out_rubros.append(
        {
            "date": month_date,
            "type": "income",
            "category": "__meta__",
            "amount": meta["venta"] or 0,
            "description": json.dumps({"meta": True, **meta}, ensure_ascii=False),
            "source_file": SOURCE_RUBRO,
        }
    )

    return channel_rows, out_rubros, meta


def is_entre_cuentas_note(note) -> bool:
    if note is None:
        return False
    key = norm_rubro_key(str(note))
    return "ENTRE CUENTA" in key


def clean_sem_note(note) -> str | None:
    """Normalize free-text concept note from SEM cols C/F/I; None if empty."""
    if note is None:
        return None
    text = re.sub(r"\s+", " ", str(note).strip())
    if not text:
        return None
    return text


# SEM channel layout: label_col, amount_col, note_col, canal name
SEM_CHANNEL_STREAMS = (
    (0, 1, 2, "Efectivo"),
    (3, 4, 5, "Mifel"),
    (6, 7, 8, "BBVA"),
)


def extract_entre_cuentas_otros(wb) -> dict[str, float]:
    """
    Scan SEM sheets for OTROS rows noted as transferencias 'Entre cuentas'
    and sum those channel amounts so they can be subtracted from OTROS real.
    Layout: Efectivo A/B + note C · Mifel D/E + note F · BBVA G/H + note I.
    """
    totals = {"efectivo": 0.0, "mifel": 0.0, "bbva": 0.0}
    for n in range(1, 6):
        name = f"SEM {n}"
        if name not in wb.sheetnames:
            continue
        ws = wb[name]
        for row in ws.iter_rows(max_row=120, max_col=10, values_only=True):
            cells = list(row) + [None] * 10
            if norm_rubro_key(str(cells[0] or "")) == "OTROS" and is_entre_cuentas_note(cells[2]):
                totals["efectivo"] += money_or_0(cells[1])
            if norm_rubro_key(str(cells[3] or "")) == "OTROS" and is_entre_cuentas_note(cells[5]):
                totals["mifel"] += money_or_0(cells[4])
            if norm_rubro_key(str(cells[6] or "")) == "OTROS" and is_entre_cuentas_note(cells[8]):
                totals["bbva"] += money_or_0(cells[7])
    return totals


def extract_sem_detalle(wb, year: int, month: int) -> list[dict]:
    """
    Per-week rubro × canal lines from SEM sheets, including free-text notes
    (cols C/F/I) used as concepts in the Real drill-down (e.g. huerta, galacticos).

    Skips parent section headers, footer labels, zero/empty noise, and OTROS
    'Entre cuentas' transfers (already excluded from OTROS real).
    """
    month_date = f"{year:04d}-{month:02d}-01"
    records: list[dict] = []

    for n in range(1, 6):
        sheet_name = f"SEM {n}"
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        last_parent: str | None = None

        for row_idx, row in enumerate(
            ws.iter_rows(max_row=120, max_col=10, values_only=True)
        ):
            if row_idx == 0:
                continue
            cells = list(row) + [None] * 10

            # Track parent sections from Efectivo column order (same as TOTAL)
            e_raw = cells[0]
            if e_raw:
                e_name = norm_cat(str(e_raw))
                e_key = norm_rubro_key(e_name)
                if e_key in PARENT_DISPLAY:
                    last_parent = PARENT_DISPLAY[e_key]
                elif e_key in SKIP_NAMES or e_name.upper() in SKIP_NAMES:
                    pass
                else:
                    parent_hint, last_parent = detect_parent(
                        e_name, row_idx, last_parent
                    )
                    # detect_parent returns parent for this row; keep last_parent
                    _ = parent_hint

            for label_col, amount_col, note_col, canal in SEM_CHANNEL_STREAMS:
                raw_label = cells[label_col]
                if not raw_label:
                    continue
                rubro_raw = norm_cat(str(raw_label))
                key = norm_rubro_key(rubro_raw)
                upper = rubro_raw.upper()
                if upper in SKIP_NAMES or key in SKIP_NAMES:
                    continue
                if key in PARENT_DISPLAY:
                    continue

                amount = money_or_0(cells[amount_col])
                note = clean_sem_note(cells[note_col])
                if amount == 0 and not note:
                    continue

                # Exclude OTROS entre-cuentas (not part of Real)
                if key == "OTROS" and is_entre_cuentas_note(note):
                    continue

                parent, _ = detect_parent(rubro_raw, row_idx, last_parent)
                # Prefer section last_parent for known children when detect returns None
                if parent is None and last_parent:
                    last_key = parent_section_key(last_parent)
                    if last_key == "INSUMOS DE COCINA" and (
                        upper in COCINA_CHILDREN
                        or upper.startswith("PANES")
                        or key == "CARBON"
                    ):
                        parent = PARENT_COCINA
                    elif last_key == "INSUMOS DE BARRA" and (
                        upper in BARRA_CHILDREN or upper in {"CAFE", "CAFÉ"}
                    ):
                        parent = PARENT_BARRA
                    elif last_key == "SERVICIOS" and (
                        upper in SERVICIOS_CHILDREN or key in SERVICIOS_CHILDREN_KEYS
                    ):
                        parent = PARENT_SERVICIOS

                rubro = canonicalize_rubro_name(rubro_raw)
                payload = {
                    "week": n,
                    "rubro": rubro,
                    "parent": parent,
                    "canal": canal,
                    "amount": amount,
                    "note": note,
                }
                records.append(
                    {
                        "date": month_date,
                        "type": "expense",
                        "category": f"SEM {n} · {rubro}",
                        "amount": amount,
                        "description": json.dumps(payload, ensure_ascii=False),
                        "source_file": SOURCE_DETALLE,
                    }
                )

    return records


def apply_entre_cuentas_correction(
    channel_rows: list[dict], rubro_rows: list[dict], entre: dict[str, float]
) -> None:
    """Subtract entre-cuentas amounts from OTROS in channel + rubro records."""
    sub_e = float(entre.get("efectivo") or 0)
    sub_m = float(entre.get("mifel") or 0)
    sub_b = float(entre.get("bbva") or 0)
    if not (sub_e or sub_m or sub_b):
        return

    for ch in channel_rows:
        cat = str(ch.get("category") or "")
        if cat == "Efectivo: OTROS" and sub_e:
            ch["amount"] = max(0.0, float(ch["amount"]) - sub_e)
        elif cat == "Mifel: OTROS" and sub_m:
            ch["amount"] = max(0.0, float(ch["amount"]) - sub_m)
        elif cat == "BBVA: OTROS" and sub_b:
            ch["amount"] = max(0.0, float(ch["amount"]) - sub_b)

    channel_rows[:] = [c for c in channel_rows if float(c.get("amount") or 0) != 0]

    for r in rubro_rows:
        if r.get("category") == "__meta__":
            continue
        try:
            payload = json.loads(r["description"])
        except Exception:
            continue
        if norm_rubro_key(str(payload.get("rubro") or r.get("category") or "")) != "OTROS":
            continue
        if payload.get("parent"):
            continue
        payload["efectivo"] = max(0.0, float(payload.get("efectivo") or 0) - sub_e)
        payload["mifel"] = max(0.0, float(payload.get("mifel") or 0) - sub_m)
        payload["bbva"] = max(0.0, float(payload.get("bbva") or 0) - sub_b)
        payload["real"] = (
            float(payload["efectivo"]) + float(payload["mifel"]) + float(payload["bbva"])
        )
        r["description"] = json.dumps(payload, ensure_ascii=False)
        r["amount"] = payload["real"]
        print(
            f"OTROS - entre cuentas: efe={sub_e} mifel={sub_m} bbva={sub_b} "
            f"-> real={payload['real']}"
        )


def read_anticipos_notes(path: Path) -> dict[tuple[str, int], str]:
    """Read Excel cell comments (and note-like values) for ANTICIPOS SEM rows on TOTAL.

    Returns {(bank, week): note_text}. Bank is MIFEL or BBVA.
    Comments require a non-read-only workbook open (data_only=False).
    """
    notes: dict[tuple[str, int], str] = {}
    try:
        wb = load_workbook(path, read_only=False, data_only=False)
    except Exception:
        return notes
    if "TOTAL" not in wb.sheetnames:
        wb.close()
        return notes
    ws = wb["TOTAL"]
    in_bbva = False
    for row_idx in range(1, 46):
        # Col O (15) marks BBVA block
        o_val = str(ws.cell(row_idx, 15).value or "").strip().upper()
        if o_val == "BBVA":
            in_bbva = True
        lab = str(ws.cell(row_idx, 12).value or "").strip().upper()  # L
        ant = re.match(r"ANTICIPOS SEM\s*(\d+)", lab)
        if not ant:
            continue
        w = int(ant.group(1))
        bank = "BBVA" if in_bbva else "MIFEL"
        parts: list[str] = []
        # Comments on label (L) and entrada amount (M)
        for col in (12, 13):
            cell = ws.cell(row_idx, col)
            if cell.comment and cell.comment.text:
                parts.append(str(cell.comment.text))
        # Some workbooks put the note as a neighboring cell value (col Q/R)
        for col in (17, 18, 19, 20):
            extra = ws.cell(row_idx, col).value
            if extra is not None and str(extra).strip():
                parts.append(str(extra))
        text = " ".join(parts).strip()
        if text:
            notes[(bank, w)] = text
    wb.close()
    return notes


def classify_anticipo_tipo(bank: str, note: str | None) -> str:
    """ventas | entre_cuentas | otro — entre cuentas solo MIFEL↔BBVA con nota."""
    if not note:
        return "otro"
    if bank not in ("MIFEL", "BBVA"):
        return "otro"
    if is_entre_cuentas_note(note):
        return "entre_cuentas"
    return "otro"


# Panel Resumen semanal de movimientos en TOTAL (cols U–Z).
# U = etiqueta · V–Z = semanas 1–5
RESUMEN_PANEL_LABELS = {
    "inicial": "inicial",
    "ingresos": "ingresos",
    "pagos mifel": "pagos_mifel",
    "comisiones": "comisiones",
    "pagos bbva": "pagos_bbva",
    "inversiones": "inversiones",
    "suma ingreso": "suma_ingreso",
    "suma gastos": "suma_gasto",
    "total": "total",
}


def extract_total_resumen_panel(wb) -> dict[int, dict[str, float]]:
    """
    Lee TOTAL!U2:Z15 — fuente canónica del Resumen semanal de movimientos.
    Devuelve {week: {inicial, ingresos, pagos_mifel, ...}}.
    """
    if "TOTAL" not in wb.sheetnames:
        return {}
    rows = list(
        wb["TOTAL"].iter_rows(
            min_row=1, max_row=20, min_col=21, max_col=26, values_only=True
        )
    )
    if len(rows) < 3:
        return {}

    # Fila 2: (None, 1, 2, 3, 4, 5) — col U vacía, V–Z = nº semana
    header = rows[1]
    week_indexes: list[tuple[int, int]] = []
    for i, v in enumerate(header):
        if i == 0:
            continue
        try:
            w = int(float(v))
        except (TypeError, ValueError):
            continue
        if 1 <= w <= 6:
            week_indexes.append((w, i))
    if not week_indexes:
        return {}

    by_week: dict[int, dict[str, float]] = {w: {"week": w} for w, _ in week_indexes}
    for row in rows[2:]:
        label = str(row[0] or "").strip().lower()
        key = RESUMEN_PANEL_LABELS.get(label)
        if not key:
            continue
        for w, idx in week_indexes:
            by_week[w][key] = money_or_0(row[idx] if idx < len(row) else None)

    # Solo semanas con al menos una fila del panel
    return {
        w: payload
        for w, payload in by_week.items()
        if any(k != "week" for k in payload)
    }


def extract_week_bank_components(
    wb, year: int, month: int, anticipos_notes: dict[tuple[str, int], str] | None = None
) -> tuple[list[dict], list[dict]]:
    """Build per-week bank roll-forward + per-bank ingreso rows.

    presupuesto_semana: panel TOTAL!U:Z (resumen semanal de movimientos).
    presupuesto_ingreso: agregados manuales TOTAL (ventas M+N + anticipos).

    presupuesto_ingreso emits separate rows:
      - tipo=ventas          (SEM n ventas M+N)
      - tipo=entre_cuentas   (ANTICIPOS with Excel note «Entre cuentas», MIFEL↔BBVA)
      - tipo=otro            (other anticipos entradas)
    """
    month_date = f"{year:04d}-{month:02d}-01"
    notes = anticipos_notes or {}
    if "TOTAL" not in wb.sheetnames:
        return [], []

    resumen_panel = extract_total_resumen_panel(wb)

    total_rows = [
        list(r) + [None] * 20
        for r in wb["TOTAL"].iter_rows(max_row=45, max_col=20, values_only=True)
    ]

    # Mifel inicial (fila 2, col M) / BBVA inicial (bloque inferior)
    mifel_inicial = money_or_0(total_rows[1][12]) if len(total_rows) > 1 else 0.0
    bbva_inicial = 0.0
    for i, cells in enumerate(total_rows):
        if str(cells[14] or "").strip().upper() != "BBVA":
            continue
        for j in range(i + 1, min(i + 6, len(total_rows))):
            val = as_amount(total_rows[j][12])
            if val is not None:
                bbva_inicial = float(val)
                break
        break
    # Weekly ventas/comisiones Mifel (rows where L=SEM n around row 5+)
    mifel_weeks: dict[int, dict] = {}
    bbva_weeks: dict[int, dict] = {}
    # Anticipos: col M = entradas (van en ingresos), col N = salidas (fila inversiones)
    mifel_inv_in: dict[int, float] = {}
    mifel_inv_out: dict[int, float] = {}
    bbva_inv_in: dict[int, float] = {}
    bbva_inv_out: dict[int, float] = {}

    in_bbva = False
    for idx, cells in enumerate(total_rows):
        if str(cells[14] or "").strip().upper() == "BBVA":
            in_bbva = True
        lab = str(cells[11] or "").strip().upper()
        m = re.match(r"SEM\s*(\d+)$", lab)
        if m:
            w = int(m.group(1))
            entry = {
                "ventas": money_or_0(cells[12]) + money_or_0(cells[13]),  # M + N
                "comisiones": money_or_0(cells[14]) + money_or_0(cells[15]),  # O + P
            }
            if in_bbva:
                bbva_weeks[w] = entry
            else:
                mifel_weeks[w] = entry

        ant = re.match(r"ANTICIPOS SEM\s*(\d+)", lab)
        if ant:
            w = int(ant.group(1))
            entrada = money_or_0(cells[12])  # M
            salida = money_or_0(cells[13])  # N
            if in_bbva:
                bbva_inv_in[w] = bbva_inv_in.get(w, 0) + entrada
                bbva_inv_out[w] = bbva_inv_out.get(w, 0) + salida
            else:
                mifel_inv_in[w] = mifel_inv_in.get(w, 0) + entrada
                mifel_inv_out[w] = mifel_inv_out.get(w, 0) + salida

    # SEM sheet bank pagos — fallback si el panel U:Z no existe
    week_pagos: dict[int, dict] = {}
    for n in range(1, 6):
        name = f"SEM {n}"
        if name not in wb.sheetnames:
            continue
        ws = wb[name]
        r1 = list(next(ws.iter_rows(min_row=1, max_row=1, max_col=8, values_only=True))) + [
            None
        ] * 8
        week_pagos[n] = {
            "pagos_mifel": money_or_0(r1[4]),
            "pagos_bbva": money_or_0(r1[7]),
        }

    weeks = sorted(set(resumen_panel.keys()) | set(week_pagos.keys()) | set(mifel_weeks.keys()) | set(bbva_weeks.keys()))
    if not weeks:
        return [], []

    records: list[dict] = []
    ingreso_records: list[dict] = []

    # ── presupuesto_semana: panel TOTAL!U:Z (canónico) ─────────────────────
    if resumen_panel:
        for w in sorted(resumen_panel.keys()):
            panel = resumen_panel[w]
            mv = mifel_weeks.get(w, {"ventas": 0.0, "comisiones": 0.0})
            bv = bbva_weeks.get(w, {"ventas": 0.0, "comisiones": 0.0})
            ingresos_mifel = mv["ventas"] + mifel_inv_in.get(w, 0.0)
            ingresos_bbva = bv["ventas"] + bbva_inv_in.get(w, 0.0)
            inicial = float(panel.get("inicial") or 0)
            ingresos = float(panel.get("ingresos") or 0)
            pagos_mifel = float(panel.get("pagos_mifel") or 0)
            comisiones = float(panel.get("comisiones") or 0)
            pagos_bbva = float(panel.get("pagos_bbva") or 0)
            inversiones = float(panel.get("inversiones") or 0)
            suma_ingreso = float(
                panel.get("suma_ingreso")
                if panel.get("suma_ingreso") is not None
                else inicial + ingresos
            )
            suma_gasto = float(
                panel.get("suma_gasto")
                if panel.get("suma_gasto") is not None
                else pagos_mifel + comisiones + pagos_bbva + inversiones
            )
            total = float(
                panel.get("total")
                if panel.get("total") is not None
                else suma_ingreso - suma_gasto
            )
            payload = {
                "week": w,
                "inicial": inicial,
                "ingresos": ingresos,
                "ingresos_mifel": ingresos_mifel,
                "ingresos_bbva": ingresos_bbva,
                "pagos_mifel": pagos_mifel,
                "comisiones": comisiones,
                "pagos_bbva": pagos_bbva,
                "inversiones": inversiones,
                "suma_ingreso": suma_ingreso,
                "suma_gasto": suma_gasto,
                "total": total,
                "source": "total_UZ",
            }
            records.append(
                {
                    "date": month_date,
                    "type": "expense",
                    "category": f"Semana {w}",
                    "amount": total,
                    "description": json.dumps(payload, ensure_ascii=False),
                    "source_file": SOURCE_SEMANA,
                }
            )
    else:
        # Fallback legacy: reconstruir desde SEM + bloques Mifel/BBVA
        inicial = mifel_inicial + bbva_inicial
        for w in weeks:
            mp = week_pagos.get(w, {"pagos_mifel": 0.0, "pagos_bbva": 0.0})
            mv = mifel_weeks.get(w, {"ventas": 0.0, "comisiones": 0.0})
            bv = bbva_weeks.get(w, {"ventas": 0.0, "comisiones": 0.0})
            ingresos_mifel = mv["ventas"] + mifel_inv_in.get(w, 0.0)
            ingresos_bbva = bv["ventas"] + bbva_inv_in.get(w, 0.0)
            ingresos = ingresos_mifel + ingresos_bbva
            comisiones = mv["comisiones"] + bv["comisiones"]
            inversiones = mifel_inv_out.get(w, 0.0) + bbva_inv_out.get(w, 0.0)
            pagos_mifel = mp["pagos_mifel"]
            pagos_bbva = mp["pagos_bbva"]
            suma_ingreso = inicial + ingresos
            suma_gasto = pagos_mifel + comisiones + pagos_bbva + inversiones
            total = suma_ingreso - suma_gasto
            payload = {
                "week": w,
                "inicial": inicial,
                "ingresos": ingresos,
                "ingresos_mifel": ingresos_mifel,
                "ingresos_bbva": ingresos_bbva,
                "pagos_mifel": pagos_mifel,
                "comisiones": comisiones,
                "pagos_bbva": pagos_bbva,
                "inversiones": inversiones,
                "suma_ingreso": suma_ingreso,
                "suma_gasto": suma_gasto,
                "total": total,
                "source": "legacy_sem",
            }
            records.append(
                {
                    "date": month_date,
                    "type": "expense",
                    "category": f"Semana {w}",
                    "amount": total,
                    "description": json.dumps(payload, ensure_ascii=False),
                    "source_file": SOURCE_SEMANA,
                }
            )
            inicial = total

    # ── presupuesto_ingreso: detalle por banco (TOTAL bloques SEM / anticipos) ─
    for w in weeks:
        mv = mifel_weeks.get(w, {"ventas": 0.0, "comisiones": 0.0})
        bv = bbva_weeks.get(w, {"ventas": 0.0, "comisiones": 0.0})
        week_monday = monday_of_month_sem(year, month, w)
        bank_parts = (
            (
                "MIFEL",
                mv["ventas"],
                mifel_inv_in.get(w, 0.0),
            ),
            (
                "BBVA",
                bv["ventas"],
                bbva_inv_in.get(w, 0.0),
            ),
        )
        for bank, ventas_amt, anticipo_amt in bank_parts:
            if ventas_amt > 0:
                ingreso_payload = {
                    "bank": bank,
                    "week": w,
                    "year": year,
                    "month": month,
                    "abono": ventas_amt,
                    "ventas": ventas_amt,
                    "anticipos_entrada": 0.0,
                    "tipo": "ventas",
                    "fecha": week_monday.isoformat(),
                    "descripcion": f"Ventas {bank} · SEM {w}",
                    "source": "presupuesto_excel",
                }
                ingreso_records.append(
                    {
                        "date": month_date,
                        "type": "income",
                        "category": f"Ventas {bank} SEM {w}",
                        "amount": ventas_amt,
                        "description": json.dumps(ingreso_payload, ensure_ascii=False),
                        "source_file": SOURCE_INGRESO,
                    }
                )
            if anticipo_amt > 0:
                note = notes.get((bank, w))
                tipo = classify_anticipo_tipo(bank, note)
                if tipo == "entre_cuentas":
                    desc = f"Entre cuentas {bank} · SEM {w}"
                    cat = f"Entre cuentas {bank} SEM {w}"
                else:
                    desc = f"Anticipo {bank} · SEM {w}"
                    cat = f"Anticipo {bank} SEM {w}"
                ingreso_payload = {
                    "bank": bank,
                    "week": w,
                    "year": year,
                    "month": month,
                    "abono": anticipo_amt,
                    "ventas": 0.0,
                    "anticipos_entrada": anticipo_amt,
                    "tipo": tipo,
                    "nota": clean_sem_note(note),
                    "fecha": week_monday.isoformat(),
                    "descripcion": desc,
                    "source": "presupuesto_excel",
                }
                ingreso_records.append(
                    {
                        "date": month_date,
                        "type": "income",
                        "category": cat,
                        "amount": anticipo_amt,
                        "description": json.dumps(ingreso_payload, ensure_ascii=False),
                        "source_file": SOURCE_INGRESO,
                    }
                )

    return records, ingreso_records


def extract_from_workbook(
    path: Path,
) -> tuple[list[dict], list[dict], list[dict], list[dict], list[dict], list[dict]]:
    parsed = parse_month_year(path.name)
    if not parsed:
        raise ValueError(f"No se pudo inferir mes/año de: {path.name}")
    year, month = parsed
    month_label = path.stem.strip()

    anticipos_notes = read_anticipos_notes(path)

    wb = load_workbook(path, read_only=True, data_only=True)
    if "TOTAL" not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Sin hoja TOTAL en {path.name}. Hojas: {wb.sheetnames}")

    total_rows = [
        list(r)
        for r in wb["TOTAL"].iter_rows(max_row=60, max_col=20, values_only=True)
    ]

    saldos = extract_saldos_from_total(total_rows, year, month, month_label)
    channel_rows, rubro_rows, _meta = extract_rubros(total_rows, year, month, month_label)
    week_rows, ingreso_rows = extract_week_bank_components(
        wb, year, month, anticipos_notes
    )
    detalle_rows = extract_sem_detalle(wb, year, month)
    entre = extract_entre_cuentas_otros(wb)
    apply_entre_cuentas_correction(channel_rows, rubro_rows, entre)
    wb.close()
    return channel_rows, saldos, rubro_rows, week_rows, detalle_rows, ingreso_rows


def find_budget_files(folder: Path) -> list[Path]:
    return sorted(
        p
        for p in folder.glob("PRESUPUESTO*.xlsx")
        if "FLUJO" not in p.name.upper()
    )


def find_all_budget_files(folders: list[Path]) -> list[Path]:
    files: list[Path] = []
    for folder in folders:
        if not folder.exists():
            print(f"SKIP carpeta (no existe): {folder}")
            continue
        found = find_budget_files(folder)
        print(f"{folder.name}: {len(found)} archivos")
        files.extend(found)
    return files


def chunked(items: list[dict], size: int = 200):
    for i in range(0, len(items), size):
        yield items[i : i + size]


def main() -> None:
    parser = argparse.ArgumentParser(description="Ingesta presupuesto mensual → Supabase")
    parser.add_argument(
        "--folder",
        type=Path,
        default=None,
        help="Una sola carpeta. Si se omite, ingiere 2022–2026.",
    )
    parser.add_argument("--file", type=Path, default=None)
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument(
        "--full",
        action="store_true",
        help="Wipe + reload de todos los años (2022–2026). Por defecto solo ~30 días.",
    )
    parser.add_argument(
        "--recent-days",
        type=int,
        default=40,
        help="Meses a refrescar (ventana en días). Default 40. Ignorado con --full.",
    )
    args = parser.parse_args()

    recent_months: set[tuple[int, int]] | None = None
    if not args.full and not args.file:
        recent_months = months_in_recent_window(args.recent_days)
        print(
            f"Modo incremental (~{args.recent_days}d): meses "
            + ", ".join(f"{y}-{m:02d}" for y, m in sorted(recent_months))
        )

    if args.file:
        files = [args.file]
    elif args.folder:
        files = find_budget_files(args.folder)
    else:
        files = find_all_budget_files(YEAR_FOLDERS)

    if recent_months is not None:
        filtered: list[Path] = []
        for path in files:
            parsed = parse_month_year(path.name)
            if parsed and parsed in recent_months:
                filtered.append(path)
            elif parsed:
                continue
            else:
                print(f"SKIP (mes no parseable): {path.name}")
        files = filtered

    if not files:
        raise SystemExit("No hay archivos PRESUPUESTO*.xlsx para ingerir")

    all_channel: list[dict] = []
    all_saldos: list[dict] = []
    all_rubros: list[dict] = []
    all_weeks: list[dict] = []
    all_detalle: list[dict] = []
    all_ingresos: list[dict] = []

    for path in files:
        if not path.exists():
            print(f"SKIP (no existe): {path}")
            continue
        try:
            channels, saldos, rubros, weeks, detalle, ingresos = extract_from_workbook(
                path
            )
        except Exception as exc:
            print(f"ERROR {path.name}: {exc}")
            continue
        print(
            f"{path.name}: canales={len(channels)} rubros={len(rubros)} "
            f"semanas={len(weeks)} saldos={len(saldos)} detalle={len(detalle)} "
            f"ingresos_banco={len(ingresos)}"
        )
        all_channel.extend(channels)
        all_saldos.extend(saldos)
        all_rubros.extend(rubros)
        all_weeks.extend(weeks)
        all_detalle.extend(detalle)
        all_ingresos.extend(ingresos)

    combined = (
        all_channel
        + all_saldos
        + all_rubros
        + all_weeks
        + all_detalle
        + all_ingresos
    )
    print(f"TOTAL registros: {len(combined)}")

    if args.dry_run:
        print("Dry-run: no se escribió nada.")
        if all_rubros:
            print("Ejemplo rubro:", all_rubros[0])
        if all_weeks:
            print("Ejemplo semana:", all_weeks[0])
        if all_ingresos:
            print("Ejemplo ingreso banco:", all_ingresos[0])
        if all_detalle:
            print("Ejemplo detalle:", all_detalle[0])
            # Show Mantenimiento sample if present
            mant = [
                d
                for d in all_detalle
                if "MANTENIMIENTO" in norm_rubro_key(
                    json.loads(d["description"]).get("rubro", "")
                )
            ]
            if mant:
                print(f"Detalle Mantenimiento ({len(mant)} líneas):")
                for d in mant[:8]:
                    print(" ", d["description"])
        # Dedup sanity: count categories per month
        from collections import Counter

        cats = Counter()
        for r in all_rubros:
            if r["category"] == "__meta__":
                continue
            try:
                payload = json.loads(r["description"])
                parent = payload.get("parent") or ""
            except Exception:
                parent = ""
            cats[f"{r['date']}:{parent}:{r['category']}"] += 1
        dups = {k: v for k, v in cats.items() if v > 1}
        print(f"Rubros duplicados post-merge: {len(dups)}")
        for k, v in list(dups.items())[:10]:
            print(f"  {k} x{v}")
        return

    url = os.environ.get("SUPABASE_URL") or os.environ.get("NEXT_PUBLIC_SUPABASE_URL")
    key = (
        os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
        or os.environ.get("SUPABASE_ANON_KEY")
        or os.environ.get("NEXT_PUBLIC_SUPABASE_ANON_KEY")
    )
    if not url or not key:
        raise SystemExit("Faltan credenciales Supabase en .env / .env.local")

    supabase = create_client(url, key)

    if args.full or args.file:
        for src in ALL_SOURCES:
            if args.file:
                # Single-file: delete only that month's rows for each source
                parsed = parse_month_year(args.file.name)
                if parsed:
                    y, m = parsed
                    month_date = f"{y:04d}-{m:02d}-01"
                    (
                        supabase.table("financial_records")
                        .delete()
                        .eq("source_file", src)
                        .eq("date", month_date)
                        .execute()
                    )
                    print(f"Limpieza {src} {month_date}: OK")
                else:
                    supabase.table("financial_records").delete().eq("source_file", src).execute()
                    print(f"Limpieza {src}: OK")
            else:
                supabase.table("financial_records").delete().eq("source_file", src).execute()
                print(f"Limpieza {src}: OK")
    else:
        # Incremental: delete only months in the window
        assert recent_months is not None
        for y, m in sorted(recent_months):
            month_date = f"{y:04d}-{m:02d}-01"
            for src in ALL_SOURCES:
                (
                    supabase.table("financial_records")
                    .delete()
                    .eq("source_file", src)
                    .eq("date", month_date)
                    .execute()
                )
            print(f"Limpieza meses {month_date}: OK")

    inserted = 0
    for batch in chunked(combined, 200):
        result = supabase.table("financial_records").insert(batch).execute()
        inserted += len(result.data or [])

    print(f"Insertados: {inserted}")


if __name__ == "__main__":
    main()
