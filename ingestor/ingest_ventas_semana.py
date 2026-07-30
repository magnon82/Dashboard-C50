"""
Ingestor: Acumulado ventas x semana.xlsx → financial_records.

Una fila por concepto (Venta WI / Eventos) por semana.
source_file = ventas_semana
"""

from __future__ import annotations

import argparse
import os
from datetime import date, timedelta
from pathlib import Path

from dotenv import load_dotenv
from openpyxl import load_workbook
from supabase import create_client

load_dotenv()

SOURCE_FILE = "ventas_semana"
DEFAULT_PATH = Path(r"I:\Mi unidad\Administración\Controles\Acumulado ventas x semana.xlsx")

MONTHS = {
    "enero": 1,
    "febrero": 2,
    "marzo": 3,
    "abril": 4,
    "mayo": 5,
    "junio": 6,
    "julio": 7,
    "agosto": 8,
    "septiembre": 9,
    "octubre": 10,
    "noviembre": 11,
    "diciembre": 12,
}


def week_anchor_date(year: int, week: int) -> str:
    """Lunes de la semana N según Acumulado (1er lunes en/después del 1 ene)."""
    jan1 = date(year, 1, 1)
    while jan1.weekday() != 0:
        jan1 += timedelta(days=1)
    monday = jan1 + timedelta(weeks=max(week, 1) - 1)
    return monday.isoformat()


def as_amount(value) -> float:
    if value is None or isinstance(value, str):
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def month_num(name) -> int | None:
    if not name or not isinstance(name, str):
        return None
    return MONTHS.get(name.strip().lower())


def extract_year(ws, year: int) -> list[dict]:
    records: list[dict] = []
    current_month_name = None
    current_month = 1

    for row in ws.iter_rows(min_row=2, max_col=7, values_only=True):
        cells = list(row) + [None] * 7
        mes, semana, _sub, _desc, venta_wi, eventos, total = cells[:7]

        if semana is None:
            continue
        if isinstance(semana, str) and "total" in semana.lower():
            continue
        try:
            week = int(semana)
        except (TypeError, ValueError):
            continue

        if mes:
            current_month_name = str(mes).strip()
            current_month = month_num(current_month_name) or current_month

        wi = as_amount(venta_wi)
        ev = as_amount(eventos)
        tot = as_amount(total)

        # Semanas vacías / solo meta
        if wi == 0 and ev == 0 and tot == 0:
            continue

        # Si TOTAL viene y WI/Eventos no, usar TOTAL como Venta WI
        if tot > 0 and wi == 0 and ev == 0:
            wi = tot

        month_label = current_month_name or f"Mes {current_month}"
        anchor = week_anchor_date(year, week)
        base_desc = f"{year} Semana {week} · {month_label}"

        if tot <= 0:
            tot = wi + ev

        if tot > 0:
            records.append(
                {
                    "date": anchor,
                    "type": "income",
                    "category": "TOTAL",
                    "amount": tot,
                    "description": base_desc,
                    "source_file": SOURCE_FILE,
                }
            )
        if wi > 0:
            records.append(
                {
                    "date": anchor,
                    "type": "income",
                    "category": "Venta WI",
                    "amount": wi,
                    "description": base_desc,
                    "source_file": SOURCE_FILE,
                }
            )
        if ev > 0:
            records.append(
                {
                    "date": anchor,
                    "type": "income",
                    "category": "Eventos",
                    "amount": ev,
                    "description": base_desc,
                    "source_file": SOURCE_FILE,
                }
            )

    return records


def extract_all(path: Path, years: list[int] | None) -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    records: list[dict] = []
    available = [s for s in wb.sheetnames if s.isdigit()]
    target = [str(y) for y in years] if years else available

    for name in target:
        if name not in wb.sheetnames:
            print(f"SKIP hoja {name} (no existe)")
            continue
        rows = extract_year(wb[name], int(name))
        total = sum(r["amount"] for r in rows)
        print(f"{name}: {len(rows)} filas, ${total:,.2f}")
        records.extend(rows)

    wb.close()
    return records


def chunked(items: list[dict], size: int = 200):
    for i in range(0, len(items), size):
        yield items[i : i + size]


def main() -> None:
    parser = argparse.ArgumentParser(description="Ingesta ventas por semana → Supabase")
    parser.add_argument("--file", type=Path, default=DEFAULT_PATH)
    parser.add_argument(
        "--years",
        type=int,
        nargs="*",
        default=None,
        help="Años a ingerir (default: todas las hojas numéricas)",
    )
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    if not args.file.exists():
        raise SystemExit(f"No se encontró: {args.file}")

    records = extract_all(args.file, args.years)
    print(f"TOTAL: {len(records)} registros, ${sum(r['amount'] for r in records):,.2f}")

    if args.dry_run:
        print("Dry-run: no se escribió nada.")
        if records:
            print("Ejemplo:", records[0])
        return

    url = os.environ.get("SUPABASE_URL")
    key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
    if not url or not key:
        raise SystemExit("Faltan SUPABASE_URL o SUPABASE_SERVICE_ROLE_KEY en .env")

    supabase = create_client(url, key)
    supabase.table("financial_records").delete().eq("source_file", SOURCE_FILE).execute()
    print(f"Limpieza previa source_file={SOURCE_FILE}: OK")

    inserted = 0
    for batch in chunked(records, 200):
        result = supabase.table("financial_records").insert(batch).execute()
        inserted += len(result.data or [])

    print(f"Insertados: {inserted}")
    print("Listo. Recarga http://localhost:3000")


if __name__ == "__main__":
    main()
