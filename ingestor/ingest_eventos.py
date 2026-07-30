"""
Ingestor: Eventos → financial_records.

Fuentes:
  1. Google Sheets EVENTOS C50 {año} — pestaña Global
  2. Excel locales en I:\\Mi unidad\\Eventos con pestaña Global

Monto evento = VENTA + VENTA EXTRA (fecha en columna FECHA).
source_file = eventos
"""

from __future__ import annotations

import argparse
import os
import re
from datetime import datetime
from pathlib import Path

from dotenv import load_dotenv
from openpyxl import load_workbook
from supabase import create_client

from google_auth import drive_service, sheets_service

load_dotenv()

SOURCE_FILE = "eventos"
DEFAULT_EVENTOS_DIR = Path(r"I:\Mi unidad\Eventos")

MESES_ES = {
    "enero": 1,
    "febrero": 2,
    "marzo": 3,
    "abril": 4,
    "mayo": 5,
    "junio": 6,
    "julio": 7,
    "agosto": 8,
    "septiembre": 9,
    "octubre": 10,
    "noviembre": 11,
    "diciembre": 12,
}


def parse_money(value) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    if not s or s in ("-", "#DIV/0!", "#N/A"):
        return 0.0
    s = s.replace("$", "").replace(",", "").replace(" ", "")
    try:
        return float(s)
    except ValueError:
        return 0.0


def parse_spanish_date(text: str) -> str | None:
    if not text or not isinstance(text, str):
        return None
    s = text.strip().lower()
    if not s:
        return None

    # "viernes, 27 de marzo de 2026"
    m = re.search(
        r"(\d{1,2})\s+de\s+(\w+)\s+de\s+(\d{4})",
        s,
    )
    if m:
        day, month_name, year = int(m.group(1)), m.group(2), int(m.group(3))
        month = MESES_ES.get(month_name)
        if month:
            return f"{year:04d}-{month:02d}-{day:02d}"

    # datetime de Excel
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s[:10], fmt).date().isoformat()
        except ValueError:
            continue
    return None


def row_to_event(row: list, source_label: str) -> dict | None:
    """Global sheet: EVENTO, VENTA (col D), VENTA EXTRA (col E), FECHA (col I)."""
    cells = list(row) + [None] * 12
    nombre = str(cells[0] or "").strip()
    if not nombre:
        return None

    venta = parse_money(cells[3])
    extra = parse_money(cells[4])
    total = venta + extra
    if total <= 0:
        return None

    fecha = parse_spanish_date(cells[8] if len(cells) > 8 else None)
    if not fecha:
        return None

    return {
        "date": fecha,
        "type": "income",
        "category": "Eventos",
        "amount": total,
        "description": f"{nombre} · {source_label} · VENTA {venta:,.2f} + EXTRA {extra:,.2f}",
        "source_file": SOURCE_FILE,
    }


def extract_from_global_rows(rows: list[list], source_label: str) -> list[dict]:
    records: list[dict] = []
    started = False
    for row in rows:
        if not row:
            continue
        header = str(row[0] if row else "").strip().upper()
        if header == "EVENTO" or (len(row) > 8 and str(row[8]).upper().startswith("FECHA")):
            started = True
            continue
        if not started:
            continue
        ev = row_to_event(row, source_label)
        if ev:
            records.append(ev)
    return records


def extract_from_xlsx(path: Path) -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    sheet_name = next((s for s in wb.sheetnames if s.lower() == "global"), None)
    if not sheet_name:
        wb.close()
        return []
    rows = list(wb[sheet_name].iter_rows(values_only=True))
    wb.close()
    return extract_from_global_rows(rows, path.name)


def find_eventos_sheets(years: list[int] | None) -> list[tuple[int, str]]:
    drive = drive_service()
    found: list[tuple[int, str]] = []
    target_years = years or list(range(2021, datetime.now().year + 2))

    for year in target_years:
        q = (
            f"name = 'EVENTOS C50 {year}' "
            "and mimeType='application/vnd.google-apps.spreadsheet'"
        )
        res = drive.files().list(q=q, pageSize=1, fields="files(id,name)").execute()
        for f in res.get("files", []):
            found.append((year, f["id"]))
    return found


def extract_from_google_sheet(sheet_id: str, year: int) -> list[dict]:
    sheets = sheets_service()
    for tab in ("Global", "GLOBAL"):
        try:
            result = (
                sheets.spreadsheets()
                .values()
                .get(spreadsheetId=sheet_id, range=f"'{tab}'!A1:Z500")
                .execute()
            )
            rows = result.get("values", [])
            if rows:
                return extract_from_global_rows(rows, f"EVENTOS C50 {year}")
        except Exception:
            continue
    return []


def scan_local_eventos(folder: Path) -> list[dict]:
    records: list[dict] = []
    if not folder.exists():
        return records
    for path in folder.rglob("*.xlsx"):
        try:
            rows = extract_from_xlsx(path)
            if rows:
                print(f"  {path.name}: {len(rows)} eventos")
                records.extend(rows)
        except Exception as exc:
            print(f"  SKIP {path.name}: {exc}")
    return records


def chunked(items: list[dict], size: int = 200):
    for i in range(0, len(items), size):
        yield items[i : i + size]


def main() -> None:
    parser = argparse.ArgumentParser(description="Ingesta eventos → Supabase")
    parser.add_argument("--folder", type=Path, default=DEFAULT_EVENTOS_DIR)
    parser.add_argument("--years", type=int, nargs="*", default=None)
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    records: list[dict] = []

    print("Google Sheets EVENTOS C50…")
    for year, sheet_id in find_eventos_sheets(args.years):
        rows = extract_from_google_sheet(sheet_id, year)
        total = sum(r["amount"] for r in rows)
        print(f"  {year}: {len(rows)} eventos, ${total:,.2f}")
        records.extend(rows)

    print(f"Excel locales en {args.folder}…")
    records.extend(scan_local_eventos(args.folder))

    # Deduplicar por fecha + descripción + monto
    seen: set[tuple] = set()
    unique: list[dict] = []
    for r in records:
        key = (r["date"], r["description"], r["amount"])
        if key in seen:
            continue
        seen.add(key)
        unique.append(r)

    print(f"TOTAL: {len(unique)} eventos, ${sum(r['amount'] for r in unique):,.2f}")

    if args.dry_run:
        if unique:
            print("Ejemplo:", unique[0])
        return

    url = os.environ.get("SUPABASE_URL")
    key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
    if not url or not key:
        raise SystemExit("Faltan SUPABASE_URL o SUPABASE_SERVICE_ROLE_KEY en .env")

    supabase = create_client(url, key)
    supabase.table("financial_records").delete().eq("source_file", SOURCE_FILE).execute()
    print(f"Limpieza previa source_file={SOURCE_FILE}: OK")

    inserted = 0
    for batch in chunked(unique, 200):
        result = supabase.table("financial_records").insert(batch).execute()
        inserted += len(result.data or [])

    print(f"Insertados: {inserted}")
    print("Listo. Recarga http://localhost:3000")


if __name__ == "__main__":
    main()
